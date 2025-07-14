import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import matplotlib.pyplot as plt
import json
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
import boto3
import datetime
from io import BytesIO
from PIL import Image
import openpyxl
import tempfile
import re
import numpy as np
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import colorsys
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configure Streamlit to reduce automatic reruns
# Carregar flat logo via URL direta
logo_flat = 'https://www.innovatismc.com.br/wp-content/uploads/2023/12/logo-innovatis-flatico-150x150.png'
st.set_page_config(layout="wide", page_title='DASHBOARD v1.4', page_icon=logo_flat)

# Add custom CSS to improve loading experience
st.markdown("""
<style>
    /* Improve loading transitions */
    .stApp {
        transition: opacity 0.3s ease-in-out;
    }
    
    /* Custom loading animation for rerunning state */
    .stStatusWidget {
        visibility: hidden;
        height: 0;
    }
    
    /* Add a subtle background animation for loading states */
    @keyframes pulse {
        0% {background-color: rgba(255, 255, 255, 0.95);}
        50% {background-color: rgba(255, 255, 255, 0.85);}
        100% {background-color: rgba(255, 255, 255, 0.95);}
    }
    
    .element-container:has(.stAlert) {
        animation: pulse 2s infinite;
    }
    
    /* Hide the hamburger menu to further reduce unwanted interactions */
    header[data-testid="stHeader"] {
        display: none;
    }
    
    /* Hide input instructions tooltip */
    .st-emotion-cache-16idsys p,
    .st-emotion-cache-1cwn1b4 {
        display: none !important;
    }
    
    /* Hide instruction text for all input fields */
    [data-baseweb="input"] ~ p,
    [data-baseweb="input"] ~ div[class*="st-emotion-cache-"] {
        display: none !important;
    }
    
    /* Ensure no floating instructions appear */
    .stTextInput div[data-testid="stText"],
    .stNumberInput div[data-testid="stText"] {
        display: none !important;
    }
    
    /* Additional catch-all for any instruction overlays */
    div[class*="st-emotion-cache-"][aria-label*="instruction"],
    div[class*="st-emotion-cache-"][role="tooltip"] {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

# Disable automatic rerunning for selectbox and multiselect on first click
if "SALDOOK_CONFIG" not in st.session_state:
    st.session_state.SALDOOK_CONFIG = {
        "initialized": True
    }
    
    # Optional: Add any future configuration settings here

# Helper function to convert rgba CSS string to HEX for openpyxl
def rgba_to_hex(rgba_string):
    """Converts CSS rgba(r, g, b, a) string or background-color style to #AARRGGBB hex string."""
    hex_color = "FFFFFFFF" # Default white
    try:
        # Match background-color: rgba(...) or just rgba(...)
        match_bg = re.search(r'rgba?\((\d+),\s*(\d+),\s*(\d+)(?:,\s*([\d\.]+))?\)', rgba_string)
        if match_bg:
            groups = match_bg.groups()
            r, g, b = map(int, groups[:3])
            # Alpha handling: openpyxl needs FF for fully opaque.
            # CSS rgba alpha is 0-1, style strings might omit it (implicitly 1)
            # We'll assume full opacity for Excel if alpha isn't explicitly low
            # For simplicity here, we map all to FF for Excel fill.
            # If specific alpha is needed later, this part can be refined.
            hex_color = f"FF{r:02X}{g:02X}{b:02X}" # Use FF alpha for solid fill in Excel

        # Fallback for simple color names or existing hex codes if needed
        elif rgba_string.startswith('#'):
             hex_code = rgba_string.lstrip('#')
             if len(hex_code) == 6:
                 hex_color = f"FF{hex_code.upper()}"
             elif len(hex_code) == 3: # shorthand hex
                 r_hex, g_hex, b_hex = tuple(c*2 for c in hex_code)
                 hex_color = f"FF{r_hex.upper()}{g_hex.upper()}{b_hex.upper()}"
             elif len(hex_code) == 8: # Already AARRGGBB
                  hex_color = hex_code.upper()
    except Exception as e:
        print(f"Error converting color '{rgba_string}' to HEX: {e}") # Add logging for debug
        pass # Ignore errors, return default
    # Final validation
    if not re.match(r"^[A-Fa-f0-9]{8}$", hex_color):
        return "FFFFFFFF" # Return white if validation fails
    return hex_color


# Helper function to generate styled Excel file
def create_styled_excel(df, project_id_col, color_mapping, numeric_cols, currency_cols, percentage_cols, filename, drop_id_col_on_export=True):
    """
    Generates a styled Excel file with row coloring based on project ID and formatting.

    Args:
        df (pd.DataFrame): DataFrame to export.
        project_id_col (str): Column name containing the ID used for coloring.
        color_mapping (dict): Dictionary mapping project IDs to HEX color strings (e.g., 'FFFF0000' for red).
        numeric_cols (list): List of column names for general number formatting.
        currency_cols (list): List of column names for currency (R$) formatting.
        percentage_cols (list): List of column names for percentage formatting.
        filename (str): Desired filename for the download.
        drop_id_col_on_export (bool): If True, removes the project_id_col from the final export.

    Returns:
        BytesIO: In-memory buffer containing the Excel file content.
    """
    output = BytesIO()
    # Make a copy to avoid modifying the original dataframe passed by reference
    df_export = df.copy()

    # --- Column Dropping Logic ---
    df_to_write = df_export.copy()
    id_col_to_drop = None
    if drop_id_col_on_export and project_id_col and project_id_col in df_to_write.columns:
        id_col_to_drop = project_id_col

    # Find the actual ID column used for coloring (handling case variations if necessary)
    id_col_found = None
    if project_id_col and project_id_col in df_export.columns: # Check if project_id_col is not None before checking if it's in columns
        id_col_found = project_id_col
    elif project_id_col: # Only search for alternatives if an ID column was intended
        potential_cols = [col for col in df_export.columns if col.upper() == project_id_col.upper()]
        if potential_cols:
            id_col_found = potential_cols[0]
            print(f"Info: Found project ID column as '{id_col_found}' for styling (original request: '{project_id_col}').")
            if drop_id_col_on_export and id_col_to_drop is None: # If original wasn't found but alternative was
                 id_col_to_drop = id_col_found
        else:
            print(f"Warning: Project ID column '{project_id_col}' not found for Excel export styling. Coloring will be skipped.")
    # If project_id_col was None or not found, id_col_found remains None, and coloring is skipped

    # Drop the ID column from the DataFrame that will be written to Excel *if requested*
    if id_col_to_drop and id_col_to_drop in df_to_write.columns:
         df_to_write = df_to_write.drop(columns=[id_col_to_drop])
         print(f"Info: Dropped column '{id_col_to_drop}' from Excel export.")


    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write the potentially modified dataframe (without the ID column if dropped)
        df_to_write.to_excel(writer, sheet_name='Dados', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Dados']

        # Define styles
        currency_format = u'"R$" #,##0.00' # Ensure unicode for currency symbol
        percentage_format = '0.0%'
        general_num_format = '#,##0' # For integer-like numbers like Quant.
        text_format = '@' # Explicitly text

        header_font = Font(bold=True, color="FFFFFFFF") # White text
        header_fill = PatternFill(start_color="FF31AA4D", end_color="FF31AA4D", fill_type="solid") # Theme green
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False) # Wrap text off for data cells initially
        thin_border_side = Side(border_style="thin", color="FFDDDDDD") # Light gray border
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Apply header style and store column index mapping (using df_to_write columns)
        col_name_to_idx = {}
        for col_idx, value in enumerate(df_to_write.columns.values, 1):
             cell = worksheet.cell(row=1, column=col_idx)
             cell.font = header_font
             cell.fill = header_fill
             cell.alignment = header_alignment
             cell.border = thin_border
             col_name_to_idx[value] = col_idx

        # Apply row styles and number formats (iterate through original df_export for coloring keys)
        for row_idx, record in enumerate(df_export.to_dict('records'), 2): # Start from row 2 (after header)
            hex_color = "FFFFFFFF" # Default white
            # Apply coloring only if id_col_found is valid and the ID exists in the record and mapping
            if id_col_found and id_col_found in record:
                 project_id_value = record[id_col_found]
                 # Check if the project_id_value exists in the mapping before accessing
                 if project_id_value in color_mapping:
                      # Ensure color retrieved from mapping is valid hex
                      retrieved_color = color_mapping[project_id_value]
                      # Validate the HEX color format (AARRGGBB)
                      if isinstance(retrieved_color, str) and re.match(r"^[A-Fa-f0-9]{8}$", retrieved_color):
                          hex_color = retrieved_color
                      else:
                           print(f"Warning: Invalid or missing color '{retrieved_color}' found in mapping for ID '{project_id_value}'. Using default white.")
                           hex_color = "FFFFFFFF" # Ensure fallback is applied
                 # If project_id not in mapping, default to white (already set)

            # Ensure hex_color is valid before creating PatternFill
            if not (isinstance(hex_color, str) and re.match(r"^[A-Fa-f0-9]{8}$", hex_color)):
                 hex_color = "FFFFFFFF" # Final fallback check

            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid") # Use ARGB

            for col_name, value in record.items():
                # Only process columns that are actually present in the *exported* sheet
                if col_name not in col_name_to_idx: continue
                col_idx = col_name_to_idx[col_name]
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Apply fill and border first
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = cell_alignment

                # Apply number/text formatting based on column type
                # Handle potential NaNs or other non-numeric types gracefully before formatting
                if pd.isna(value):
                    cell.value = "—" # Replace NaN with dash
                    cell.number_format = text_format
                elif col_name in currency_cols:
                    # Ensure value is numeric before applying format
                    try:
                        # Attempt to convert, handling potential strings with currency symbols/separators
                        if isinstance(value, str):
                            # Clean string for conversion
                            cleaned_value = re.sub(r'[R$\s._]', '', value).replace(',', '.') # More robust cleaning
                            numeric_value = float(cleaned_value)
                        else:
                            numeric_value = float(value)

                        cell.value = numeric_value
                        cell.number_format = currency_format
                    except (ValueError, TypeError):
                        cell.value = str(value) # Keep as string if not convertible
                        cell.number_format = text_format
                elif col_name in percentage_cols:
                    try:
                        # Convert percentage written as X.Y% or X.Y to 0.XY format
                        numeric_value = 0 # Default
                        if isinstance(value, str):
                            if '%' in value:
                                numeric_value = float(value.replace('%','').replace(',','.')) / 100.0
                            else:
                                # Handle cases where percentage might be passed as string number 'X.Y'
                                numeric_value = float(value.replace(',','.')) / 100.0
                        elif isinstance(value, (int, float)):
                             # Check if value is already a decimal (e.g., 0.6) or a whole number percent (e.g., 60)
                            numeric_value = value if abs(value) <= 1 else value / 100.0

                        cell.value = numeric_value
                        cell.number_format = percentage_format
                    except (ValueError, TypeError):
                         cell.value = str(value)
                         cell.number_format = text_format
                elif col_name in numeric_cols:
                     try:
                         # Attempt int conversion first for whole numbers
                         numeric_value = int(value)
                         cell.value = numeric_value
                         cell.number_format = general_num_format
                     except (ValueError, TypeError):
                        # Try float conversion if int fails
                        try:
                            numeric_value = float(value)
                            cell.value = numeric_value
                            # Decide if float needs specific format or just general number
                            cell.number_format = '#,##0.00' # Example: format as float with 2 decimals
                        except (ValueError, TypeError):
                            # If all numeric conversions fail, treat as text
                            cell.value = str(value)
                            cell.number_format = text_format
                else:
                    # Default to text format for other columns
                    cell.value = str(value) # Ensure value is string
                    cell.number_format = text_format


        # Adjust column widths based on content (using df_to_write columns)
        for col_name, col_idx in col_name_to_idx.items():
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            # Check header length
            if len(str(col_name)) > max_length:
                 max_length = len(str(col_name))
            # Check cell content length
            for i in range(2, worksheet.max_row + 1): # Check data rows
                cell_value = worksheet.cell(row=i, column=col_idx).value
                if cell_value:
                    # Consider formatted length for numbers
                    num_format = worksheet.cell(row=i, column=col_idx).number_format
                    try:
                        if isinstance(cell_value, (int, float)):
                             # Use openpyxl's number format codes to approximate displayed length
                            if num_format == currency_format:
                                 formatted_value = f"R$ {cell_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") # Basic BRL format
                            elif num_format == percentage_format:
                                 formatted_value = f"{cell_value:.1%}".replace(".", ",")
                            elif num_format == general_num_format:
                                  formatted_value = f"{cell_value:,.0f}".replace(",", ".")
                            elif '#,##0.00' in num_format:
                                 formatted_value = f"{cell_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                            else:
                                 formatted_value = str(cell_value) # Fallback
                            cell_len = len(formatted_value)
                        else:
                             cell_len = len(str(cell_value))

                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception: # Catch any formatting errors
                         cell_len = len(str(cell_value))
                         if cell_len > max_length:
                              max_length = cell_len

            adjusted_width = (max_length + 2) * 1.1 # Add padding and slight multiplier
            # Set sensible min/max width
            worksheet.column_dimensions[column_letter].width = min(max(adjusted_width, 12), 50)


    output.seek(0)
    return output

# ---------------------------------------------------
# Standardized Color Generation Logic
# ---------------------------------------------------
def get_color_families_rgba():
    """Returns the standard dictionary of color families."""
    return {
        'blues': [
            'background-color: rgba(220, 240, 255, 0.65)',
            'background-color: rgba(200, 230, 255, 0.65)',
            'background-color: rgba(210, 220, 250, 0.65)',
        ],
        'greens': [
            'background-color: rgba(220, 255, 230, 0.65)',
            'background-color: rgba(210, 255, 210, 0.65)',
            'background-color: rgba(190, 245, 210, 0.65)',
        ],
        'yellows': [
            'background-color: rgba(255, 250, 210, 0.65)',
            'background-color: rgba(255, 240, 200, 0.65)',
            'background-color: rgba(253, 235, 200, 0.65)',
        ],
        'pinks': [
            'background-color: rgba(255, 220, 235, 0.65)',
            'background-color: rgba(255, 210, 205, 0.65)',
            'background-color: rgba(255, 200, 215, 0.65)',
        ],
        'grays': [
            'background-color: rgba(230, 240, 250, 0.65)',
            'background-color: rgba(225, 235, 245, 0.65)',
            'background-color: rgba(235, 245, 255, 0.65)',
        ],
        'oranges': [
            'background-color: rgba(255, 235, 225, 0.65)',
            'background-color: rgba(255, 225, 200, 0.65)',
            'background-color: rgba(255, 220, 190, 0.65)',
        ]
    }

def generate_project_color_map(project_ids_or_keys, style='rgba'):
    """
    Generates a color map for project IDs using a standardized, interleaved palette.

    Args:
        project_ids_or_keys (list or pd.Series): Unique identifiers for projects/groups.
        style (str): 'rgba' for CSS strings, 'hex' for Excel AARRGGBB hex codes.

    Returns:
        dict: Mapping {project_id: color_string}
    """
    color_families = get_color_families_rgba()
    all_colors_rgba = []
    max_family_len = max(len(colors) for colors in color_families.values())
    families = list(color_families.keys())

    # Interleave colors
    for i in range(max_family_len):
        for family in families:
            if i < len(color_families[family]):
                all_colors_rgba.append(color_families[family][i])

    color_map = {}
    num_colors = len(all_colors_rgba)
    unique_ids = pd.unique(project_ids_or_keys) # Ensure uniqueness

    for i, project_id in enumerate(unique_ids):
        rgba_color = all_colors_rgba[i % num_colors]
        if style == 'hex':
            color_map[project_id] = rgba_to_hex(rgba_color)
        else: # Default to rgba
             # Extract only the rgba(...) part for direct use in pandas styling
             match = re.search(r'rgba?\([^)]+\)', rgba_color)
             color_map[project_id] = match.group(0) if match else 'rgba(255, 255, 255, 0)' # fallback transparent

    return color_map

# ---------------------------------------------------
# Styling function using the standardized color map
# ---------------------------------------------------
def highlight_projects_detail(s, dataframe, id_column):
    """Applies background color based on project ID using the standard palette."""
    # Ensure dataframe has the necessary ID column
    if id_column not in dataframe.columns:
        print(f"Warning: ID column '{id_column}' not found in dataframe for styling.")
        return [''] * len(s) # Return empty styles

    # Generate the color map on the fly based on the current dataframe's IDs
    unique_ids = dataframe[id_column].unique()
    color_map_rgba = generate_project_color_map(unique_ids, style='rgba')

    # Map the row's project ID to a color CSS string
    project_id_value = s[id_column]
    color = color_map_rgba.get(project_id_value, 'rgba(255, 255, 255, 0)') # Default transparent
    return [f'background-color: {color}'] * len(s)

# ---------------------------------------------------
# Global color mapping for consistent colors across tables
# ---------------------------------------------------
def get_global_color_mapping(project_ids, style='rgba'):
    """Returns a consistent color mapping for all tables."""
    return generate_project_color_map(project_ids, style)

# Configuração da seção do Github para carregar dados do S3
# Importa a fonte Poppins do Google Fonts
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
        [class^=st-emotion] {
            font-family: 'Poppins', sans-serif !important;
        }
        body * {
            font-family: 'Poppins', sans-serif !important;
        }
    </style>
""", unsafe_allow_html=True)

# Título do aplicativo
st.title('Dashboard Financeiro (v1.4)')

# Importar arquivo de configuração
with open('config.yaml') as file:
    config = yaml.load(file, Loader=SafeLoader)

# Criar o objeto de autenticação
authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days']
)

authenticator.login()

# Verificação do status da autenticação
if st.session_state["authentication_status"]:
    authenticator.logout("Logout", "main", key="logout_sidebar")
    st.write(f"Bem-vindo, {st.session_state['name']}!")
elif st.session_state["authentication_status"] is False:
    st.error('Usuário/Senha inválido')
elif st.session_state["authentication_status"] is None:
    st.markdown("""
        <style>
        div[data-testid="stAppViewContainer"] {
            max-width: 600px;
            margin: auto;
        }
        </style>
    """, unsafe_allow_html=True)
    st.warning('Por Favor, utilize seu usuário e senha!')

# O resto do código só executa se autenticado
if st.session_state["authentication_status"]:
    # Configurar acesso ao S3
    s3 = boto3.resource(
        service_name='s3',
        region_name='us-east-2',
        aws_access_key_id=st.secrets["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=st.secrets["AWS_SECRET_ACCESS_KEY"]
    )

    # ---------------------------------------------------
    # Função para baixar a planilha Excel exportada do Google Sheets
    # ---------------------------------------------------
    @st.cache_data
    def baixar_planilha_excel():
        try:
            # Baixar a chave JSON diretamente do S3
            obj = s3.Bucket('jsoninnovatis').Object('chave2.json').get()
            creds_json = json.loads(obj['Body'].read().decode('utf-8'))
            
            # Definir os escopos para acesso ao Google Drive e leitura da planilha
            scope = [
                'https://www.googleapis.com/auth/drive',
                'https://www.googleapis.com/auth/spreadsheets.readonly'
            ]
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
            
            # Conectar via gspread para abrir a planilha e obter seu file ID
            client = gspread.authorize(creds)
            planilha = client.open("Valores a Receber Innovatis - Fundações")
            file_id = planilha.id  # Obtém o ID da planilha
            
            # Conectar à API do Google Drive para exportar a planilha como XLSX
            drive_service = build('drive', 'v3', credentials=creds)
            request = drive_service.files().export_media(
                fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            # Salva o conteúdo em um arquivo temporário com extensão .xlsx
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            with open(temp_file.name, "wb") as f:
                f.write(fh.getvalue())
            return temp_file.name
        except gspread.exceptions.SpreadsheetNotFound:
            st.error("Planilha não encontrada. Verifique se o nome está correto e se o serviço tem acesso.")
            # Fornecer lista de planilhas disponíveis
            try:
                available_sheets = [sheet.title for sheet in client.openall()]
                if available_sheets:
                    st.info(f"Planilhas disponíveis: {', '.join(available_sheets)}")
            except:
                pass
            st.stop()
        except Exception as e:
            st.error(f"Erro ao baixar a planilha: {str(e)}")
            st.stop()

    # ---------------------------------------------------
    # Baixar a planilha Excel exportada do Google Sheets
    # ---------------------------------------------------
    arquivo = baixar_planilha_excel()
    # ---------------------------------------------------
    # Pré-processamento com openpyxl para tratar células mescladas
    # ---------------------------------------------------
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    sheets_to_process = [
        sheet for sheet in wb.sheetnames 
        if sheet.strip().upper() not in ["GERAL_PROJETOS_FADEX", "TEDS SEM CONTRATO"]
    ]
    for sheet in sheets_to_process:
        ws = wb[sheet]
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value

    # Salva o workbook modificado em um novo arquivo temporário
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    arquivo = temp_file.name  # Atualiza o caminho para o arquivo tratado

    # ---------------------------------------------------
    # Continuação do seu código original para processar o Excel
    # ---------------------------------------------------

    df_raw = pd.read_excel(arquivo, sheet_name=None)  # Carrega todas as abas do Excel tratado
    # A exibição dos dados do Google Sheets foi desativada:
    # st.write("Dados consolidados do Google Sheets:")
    # st.dataframe(df_raw)


    # Formatação para exibição dos floats com duas casas decimais
    pd.options.display.float_format = '{:.2f}'.format
    # ---------------------------------------------------
    # Função para limpar valores monetários
    # ---------------------------------------------------
    def clean_numeric(x):
        try:
            if pd.isna(x) or x == "":
                return 0
            x_str = str(x)
            cleaned = re.sub(r"[^\d\.,-]", "", x_str)
            if "," in cleaned:
                cleaned = cleaned.replace(".", "").replace(",", ".")
            return float(cleaned)
        except Exception:
            return 0

    # ---------------------------------------------------
    # Função para substituir valores nulos por travessão em qualquer coluna de texto
    # ---------------------------------------------------
    def replace_none_with_dash(df):
        df_copy = df.copy()
        # Lista de todas as colunas
        for col in df_copy.columns:
            # Verificar se é uma coluna de texto
            if df_copy[col].dtype == 'object':
                df_copy[col] = df_copy[col].fillna("—")
                df_copy[col] = df_copy[col].replace(r'^\s*$', "—", regex=True)
                df_copy[col] = df_copy[col].replace('None', "—")
                df_copy[col] = df_copy[col].replace('nan', "—")
        return df_copy

    # ---------------------------------------------------
    # Função para gerar cores únicas para cada projeto baseado no HSV color space
    # ---------------------------------------------------
    def generate_unique_colors(n):
        import colorsys
        colors = []
        # Usar golden ratio para distribuir cores com máximo contraste
        golden_ratio_conjugate = 0.618033988749895
        
        # Começar de um valor aleatório para evitar cores predefinidas
        h = 0.33  # Começar na região verde-azulada
        
        for i in range(n):
            # Usar golden ratio para distribuir as cores
            h = (h + golden_ratio_conjugate) % 1.0
            
            # Usar saturação muito baixa e alto valor para cores muito suaves/translúcidas
            saturation = 0.35 + (i % 3) * 0.05  # Aumentar saturação para cores mais pronunciadas
            value = 0.98                         # Valor quase branco, mas ligeiramente reduzido
            
            rgb = colorsys.hsv_to_rgb(h, saturation, value)
            
            # Converter para rgba com transparência
            r = int(rgb[0] * 255)
            g = int(rgb[1] * 255)
            b = int(rgb[2] * 255)
            opacity = 0.4  # Aumentar opacidade para cores mais visíveis
            
            rgba_color = f"rgba({r}, {g}, {b}, {opacity})"
            colors.append(rgba_color)
        return colors

    # ---------------------------------------------------
    # Função para carregar dados de desvio
    # ---------------------------------------------------
    @st.cache_data
    def carregar_dados_desvio():
        xls = pd.ExcelFile(arquivo)
        paginas_processar = [
            pagina for pagina in xls.sheet_names 
            if pagina.strip().upper() not in ["GERAL_PROJETOS_FADEX", "TEDS SEM CONTRATO"]
        ]

        lista_dfs = []
        cols_interesse = [
            'QUANT.', 'CLIENTE', 'PROJETO', 'NOMENCLATURA DO PROJETO', 'Nº TED',
            'SECRETARIA', 'CONTRATO', # Added CONTRATO
            'VALOR DO CONTRATO', 'PREVISÃO DE VALOR DE RECEBIMENTO',
            'PREVISÃO DE DATA DE RECEBIMENTO', 'REPASSE RECEBIDO', 'DATA DE RECEBIMENTO', 'CUSTOS INCORRIDOS', 'VALOR',
            'OUTROS CORRELATOS', 'VALOR2', 'SALDO A RECEBER', 'SALDO A RECEBER DO CONTRATO'
        ]

        all_columns_found = True # Flag to track if all interest columns are present across sheets

        for pagina in paginas_processar:
            try:
                df = pd.read_excel(arquivo, sheet_name=pagina, skiprows=3)
                df.columns = df.columns.str.strip().str.upper().str.replace("  ", " ")

                # Check for missing essential columns for *this specific sheet*
                missing_essential = [col for col in ['QUANT.', 'CLIENTE', 'PROJETO', 'VALOR DO CONTRATO'] if col not in df.columns]
                if missing_essential:
                     print(f"Aviso: Pulando página '{pagina}' devido à falta de colunas essenciais: {', '.join(missing_essential)}")
                     continue # Skip this sheet if essential columns for ID are missing

                # Check which of the interest columns are actually present
                present_cols = [col for col in cols_interesse if col in df.columns]
                missing_in_sheet = [col for col in cols_interesse if col not in df.columns]
                if missing_in_sheet:
                     print(f"Aviso: Página '{pagina}' não possui as colunas: {', '.join(missing_in_sheet)}. Serão preenchidas com NA.")
                     # Add missing columns with NA to ensure consistency
                     for col in missing_in_sheet:
                         df[col] = pd.NA

                df = df[present_cols + missing_in_sheet] # Select present and add missing ones

                df = df.replace(r'^\s*$', pd.NA, regex=True)

                # Ensure QUANT. exists and filter based on it
                if 'QUANT.' in df.columns:
                     df = df.dropna(subset=["QUANT."])
                else:
                     # If QUANT. is missing entirely, we can't filter by it, proceed with caution
                     print(f"Aviso: Coluna 'QUANT.' não encontrada na página '{pagina}'. Não foi possível filtrar por ela.")


                # Define numeric columns dynamically based on what's available
                numeric_cols_to_clean = [
                    "VALOR DO CONTRATO", "REPASSE RECEBIDO", "PREVISÃO DE VALOR DE RECEBIMENTO",
                    "CUSTOS INCORRIDOS", "VALOR", "OUTROS CORRELATOS", "VALOR2",
                    "SALDO A RECEBER", "SALDO A RECEBER DO CONTRATO"
                ]
                for col in numeric_cols_to_clean:
                    if col in df.columns: # Clean only if column exists
                        df[col] = df[col].fillna(0).apply(clean_numeric)
                    else:
                        df[col] = 0 # Set default 0 if column was missing

                # Extrair fundação da página
                def extrair_fundacao(pagina):
                    if pagina.upper() == 'FUNCERN':
                        return 'FUNCERN'
                    elif 'FAPTO' in pagina.upper():
                        return 'FAPTO'
                    padrao = r'\((.*?)\)'
                    resultado = re.search(padrao, pagina)
                    if resultado:
                        return resultado.group(1).strip()
                    return pagina

                df['FUNDAÇÃO'] = extrair_fundacao(pagina)
                df['PÁGINA'] = pagina

                # Determinar o tipo com base na página ou no projeto
                df['TIPO'] = 'Projeto'
                if 'PROJETO' in df.columns: # Check if PROJETO column exists
                    produtos_mask = (
                        (pagina.upper() == 'PRODUTOS') |
                        ('PRODUTO' in pagina.upper()) |
                        (pagina == 'Produtos (FADEX)') |
                        (df['PROJETO'].astype(str).str.upper().str.contains('PRODUTO', na=False)) # Convert to string before .str
                    )
                    df.loc[produtos_mask, 'TIPO'] = 'Produto'
                else:
                    # Handle case where PROJETO column might be missing
                     print(f"Aviso: Coluna 'PROJETO' não encontrada na página '{pagina}'. Tipo padrão 'Projeto' será usado.")

                # Formatar datas if column exists
                if 'PREVISÃO DE DATA DE RECEBIMENTO' in df.columns:
                    df['PREVISÃO DE DATA DE RECEBIMENTO'] = pd.to_datetime(
                        df['PREVISÃO DE DATA DE RECEBIMENTO'], errors='coerce' # More flexible parsing initially
                    )
                    # Only format valid dates, keep NaT otherwise, then fillna
                    df['PREVISÃO DE DATA DE RECEBIMENTO'] = df['PREVISÃO DE DATA DE RECEBIMENTO'].dt.strftime('%m/%Y').fillna('A definir')
                else:
                     df['PREVISÃO DE DATA DE RECEBIMENTO'] = 'A definir' # Default if column missing


                # Limpar valores do cliente if column exists
                if 'CLIENTE' in df.columns:
                     df['CLIENTE'] = df['CLIENTE'].astype(str).replace({'': 'Não identificado', 'nan': 'Não identificado'}).fillna('Não identificado') # Ensure string type and handle nan/None
                else:
                     df['CLIENTE'] = 'Não identificado' # Default if column missing

                lista_dfs.append(df)

            except Exception as e:
                st.error(f"Erro ao processar a página '{pagina}': {e}")
                # Optionally continue to next sheet or stop
                continue # Continue processing other sheets

        if lista_dfs:
            df_consolidado = pd.concat(lista_dfs, ignore_index=True)

            # --- Create PROJETO_ID_KEY using available columns ---
            key_cols = []
            if 'QUANT.' in df_consolidado.columns:
                 key_cols.append('QUANT.')
            else:
                 print("Aviso: Coluna 'QUANT.' não encontrada para criar PROJETO_ID_KEY.")
            if 'CLIENTE' in df_consolidado.columns:
                 key_cols.append('CLIENTE')
            else:
                 print("Aviso: Coluna 'CLIENTE' não encontrada para criar PROJETO_ID_KEY.")
            if 'VALOR DO CONTRATO' in df_consolidado.columns:
                 key_cols.append('VALOR DO CONTRATO')
            else:
                 print("Aviso: Coluna 'VALOR DO CONTRATO' não encontrada para criar PROJETO_ID_KEY.")

            if len(key_cols) >= 2: # Require at least two keys for a reasonable identifier
                 df_consolidado['PROJETO_ID_KEY'] = df_consolidado.apply(
                     lambda row: '_'.join(str(row[col]).strip() for col in key_cols),
                     axis=1
                 )
                 print(f"Info: PROJETO_ID_KEY criado usando colunas: {key_cols}")
            else:
                 # Fallback: Create a less specific key or handle error
                 print("Erro Crítico: Não foi possível criar um PROJETO_ID_KEY confiável devido à falta de colunas ('CLIENTE', 'VALOR DO CONTRATO'). Usando índice como fallback.")
                 df_consolidado['PROJETO_ID_KEY'] = df_consolidado.index.astype(str) + "_fallback"
            # --- End PROJETO_ID_KEY Creation ---


            # Garantir que a coluna PÁGINA existe para evitar KeyError na seção de desvio
            if 'PÁGINA' not in df_consolidado.columns:
                # Try to derive from FUNDAÇÃO if it exists
                if 'FUNDAÇÃO' in df_consolidado.columns:
                     df_consolidado['PÁGINA'] = df_consolidado['FUNDAÇÃO'].astype(str)
                else:
                     df_consolidado['PÁGINA'] = "Página Desconhecida" # Fallback

            # Replace any remaining NaN/NaT/None in object columns with "—" before returning
            for col in df_consolidado.select_dtypes(include='object').columns:
                 df_consolidado[col] = df_consolidado[col].fillna("—")

            # Ensure numeric columns used later are present, default to 0 if not
            final_numeric_check = ['VALOR DO CONTRATO', 'REPASSE RECEBIDO', 'CUSTOS INCORRIDOS', 'VALOR', 'OUTROS CORRELATOS', 'VALOR2', 'SALDO A RECEBER', 'SALDO A RECEBER DO CONTRATO', 'PREVISÃO DE VALOR DE RECEBIMENTO']
            for col in final_numeric_check:
                if col not in df_consolidado.columns:
                     df_consolidado[col] = 0
                else:
                     # Ensure they are numeric, coercing errors to NaN then filling with 0
                     df_consolidado[col] = pd.to_numeric(df_consolidado[col], errors='coerce').fillna(0)


            # Add Project ID for Deviation section grouping (uses more columns)
            # Make sure these columns exist before using them
            factorize_cols = []
            if 'QUANT.' in df_consolidado.columns: factorize_cols.append('QUANT.')
            if 'CLIENTE' in df_consolidado.columns: factorize_cols.append('CLIENTE')
            if 'PROJETO' in df_consolidado.columns: factorize_cols.append('PROJETO')

            if len(factorize_cols) >= 2: # Need at least client and project for a meaningful ID
                df_consolidado['PROJECT ID'] = pd.factorize(
                    df_consolidado[factorize_cols].apply(lambda row: '_'.join(row.astype(str)), axis=1)
                )[0] + 1
            else:
                print("Warning: Could not create reliable 'PROJECT ID' for deviation due to missing columns. Using index.")
                df_consolidado['PROJECT ID'] = df_consolidado.index + 1

            return df_consolidado
        else:
            st.warning("Nenhuma página válida encontrada ou processada. Verifique o arquivo Excel e os nomes das abas.")
            # Return DataFrame vazio com as colunas esperadas para evitar erros posteriores
            # Include the key column as well
            empty_cols = cols_interesse + ['FUNDAÇÃO', 'PÁGINA', 'TIPO', 'PROJETO_ID_KEY', 'PROJECT ID']
            return pd.DataFrame(columns=empty_cols)

    # ---------------------------------------------------
    # Função para carregar dados processados para o dashboard (com tratamento para saldo a receber)
    # ---------------------------------------------------
    @st.cache_data
    def carregar_dados_dashboard():
        # Usar mesma base, mas aplicar tratamento para evitar duplicidade no saldo a receber
        df_base = carregar_dados_desvio().copy()
        
        # Criar chave única para cada projeto para tratar o saldo a receber
        df_base['PROJETO KEY'] = df_base.apply(
            lambda row: f"{str(row['PÁGINA']).strip()}_{str(row['QUANT.']).strip()}", 
            axis=1
        )
        
        # Manter apenas o primeiro saldo a receber de cada projeto e zerar os demais
        df_base['SALDO A RECEBER'] = df_base.groupby('PROJETO KEY')['SALDO A RECEBER'].transform(
            lambda x: [x.iloc[0]] + [0] * (len(x)-1)
        )
        
        # Remover linhas com saldo zero se desejar
        df_base = df_base[df_base["SALDO A RECEBER"] > 0]
        df_base = df_base.drop('PROJETO KEY', axis=1)
        
        # Renomear coluna de data para padronização
        df_base = df_base.rename(columns={
            'PREVISÃO DE DATA DE RECEBIMENTO': 'DATA'
        })
        
        return df_base

    @st.cache_data
    def load_logo():
        try:
            logo_obj = s3.Bucket('jsoninnovatis').Object('Logo.png').get()
            logo_data = logo_obj['Body'].read()
            return logo_data  # Return raw bytes instead of PIL Image
        except Exception as e:
            st.warning(f"Erro ao carregar logo: {e}")
            return None  # Return None for fallback

    # ---------------------------------------------------
    # Carregar dados
    # ---------------------------------------------------
    data_load_state = st.text('Carregando dados...')

    # Carregar dados para análise de desvio (sem tratamento especial para saldo a receber)
    df_desvio = carregar_dados_desvio()

    # Carregar dados para o dashboard (com tratamento para saldo a receber)
    data = carregar_dados_dashboard()

    data_load_state.text('Carregamento de dados concluído!')

    # Carregar e exibir logo
    logo_image = load_logo()
    if logo_image is not None:
        st.sidebar.image(logo_image, use_container_width=True)
    else:
        # Fallback para logo via URL direta se a logo do S3 falhar
        st.sidebar.image(logo_flat, use_container_width=True)
    
    # Definir paleta de cores para gráficos
    colors_palette = ['#a1c9f4', '#a1f4c9', '#f4d1a1', '#f4a1d8', '#c9a1f4', '#f4a1a1', '#a1f4f4', '#e6e6a1']
    
    # Estilização
    st.markdown("""
        <style>
            [data-testid="stSidebar"] * {
                font-size: 101% !important;
            }
            .st-fx { background-color: rgb(49, 170, 77); }
            .st-cx { border-bottom-color: rgb(49, 170, 77); }
            .st-cw { border-top-color: rgb(49, 170, 77); }
            .st-cv { border-right-color: rgb(49, 170, 77); }
            .st-cu { border-left-color: rgb(49, 170, 77); }
            .st-ei { background-color: #28a74500 !important; }
            .st-e2 { background: linear-gradient(to right, rgba(151, 166, 195, 0.25) 0%, rgba(151, 166, 195, 0.25) 0%, rgb(49 170 77 / 0%) 0%, rgb(49 170 77 / 0%) 100%, rgba(151, 166, 195, 0.25) 100%, rgba(151, 166, 195, 0.25) 100%); }
            .st-e3 { background: linear-gradient(to right, rgba(151, 166, 195, 0.25) 0%, rgba(151, 166, 195, 0.25) 0%, rgb(49, 170, 77) 0%, rgb(49, 170, 77) 100%, rgba(151, 166, 195, 0.25) 100%, rgba(151, 166, 195, 0.25) 100%); }
            .st-emotion-cache-1dj3ksd { background-color: #28a745 !important; }
            .st-emotion-cache-15fru4 { padding: 0.2em 0.4em; overflow-wrap: break-word; margin: 0px; border-radius: 0.25rem; background: rgb(248, 249, 251); color: rgb(9, 171, 59); font-family: "Source Code Pro", monospace; font-size: 0.75em; }
            .st-emotion-cache-1373cj4 { font-family: "Source Code Pro", monospace; font-size: 14px; color: rgb(49, 170, 77); top: -1.6em; position: absolute; white-space: nowrap; background-color: transparent; line-height: 1.6; font-weight: 400; pointer-events: none; }
            .st-fi { background-color: rgb(49, 170, 77); }
            .st-hy { background-color: rgb(49, 170, 77); }
            .st-f1 { background-color: rgb(49, 170, 77); }
            
            /* Botões com fundo branco e detalhes verdes */
            .stButton > button, .stDownloadButton > button {
                background-color: white !important;
                color: rgb(49, 170, 77) !important;
                border-color: rgb(49, 170, 77) !important;
                border-width: 1px !important;
                border-style: solid !important;
                font-weight: 500 !important;
            }
            .stButton > button:hover, .stDownloadButton > button:hover {
                background-color: rgba(49, 170, 77, 0.1) !important;
                border-color: rgb(49, 170, 77) !important;
            }
            .stButton > button:active, .stDownloadButton > button:active {
                background-color: rgba(49, 170, 77, 0.2) !important;
            }
            
            
            /* Ajuste apenas para a barra do slider e seus valores */
            [data-testid="stSlider"] [data-baseweb="slider"] div[data-testid="stThumbValue"] {
                color: rgb(49, 170, 77) !important;
                background-color: transparent !important;
                border: none !important;
            }
            
            [data-testid="stSlider"] [data-baseweb="slider-track"] {
                background-color: rgba(49, 170, 77, 0.2) !important;
            }
            
            [data-testid="stSlider"] [data-baseweb="slider-track-fill"] {
                background-color: rgb(49, 170, 77) !important;
            }
            
            [data-testid="stSlider"] [data-baseweb="slider-thumb"] {
                background-color: rgb(49, 170, 77) !important;
            }
            
            /* Valores do slider (que estão em vermelho) */
            [data-testid="stSlider"] span {
                color: rgb(49, 170, 77) !important;
            }
            
            /* Texto do slider */
            [data-testid="stSlider"] p {
                color: rgb(49, 170, 77) !important;
            }
            
            /* Multiselect e selectbox */
            .stMultiSelect > div > div > div {
                border-color: rgb(49, 170, 77) !important;
            }
            .stMultiSelect > div > div > div:hover {
                border-color: rgb(39, 150, 67) !important;
            }
            .stMultiSelect > div[data-baseweb="tag"] {
                background-color: rgb(49, 170, 77) !important;
            }
            
            /* Tabs e outros elementos de navegação */
            .stTabs [data-baseweb="tab-list"] {
                border-bottom-color: rgb(49, 170, 77) !important;
            }
            .stTabs [data-baseweb="tab"][aria-selected="true"] {
                color: rgb(49, 170, 77) !important;
                border-bottom-color: rgb(49, 170, 77) !important;
            }
            
            /* Progress bar */
            .stProgress > div > div > div > div {
                background-color: rgb(49, 170, 77) !important;
            }
            
            /* Métricas */
            .stMetric > div[data-testid="stMetricDelta"] > div {
                color: rgb(49, 170, 77) !important;
            }
            
            /* Cores de links */
            a {
                color: rgb(49, 170, 77) !important;
            }
            a:hover {
                color: rgb(39, 150, 67) !important;
            }
            
            /* Cores de foco */
            :focus {
                outline-color: rgb(49, 170, 77) !important;
            }
            
            /* Cores de seleção */
            ::selection {
                background-color: rgba(49, 170, 77, 0.3) !important;
            }
            
            /* Cores para elementos de data input */
            input[type="date"] {
                color: rgb(49, 170, 77) !important;
            }
            
            /* Cores para tooltips */
            [data-testid="stTooltipIcon"] path {
                fill: rgb(49, 170, 77) !important;
            }

            /* Corrigir cor de fundo dos campos de seleção */
            .stSelectbox > div > div,
            .stMultiSelect > div > div {
                background-color: white !important;
            }

            /* Manter apenas a borda verde */
            .stSelectbox > div > div[data-baseweb="select"] {
                border-color: #31aa4d !important;
                background-color: white !important;
            }

            /* Corrigir cor de fundo do dropdown quando aberto */
            .stSelectbox > div > div > div {
                background-color: white !important;
            }

            /* Corrigir cor de hover nas opções */
            .stSelectbox > div > div > div:hover {
                background-color: rgba(49, 170, 77, 0.1) !important;
            }

            /* Corrigir cor de fundo do slider */
            .stSlider > div > div > div {
                background-color: white !important;
            }
                
            .stSlider > div > div > div {
                background-color: #ffffff00 !important;
            }
                
            /* Corrigir cor do texto do slider */
            [data-testid="stSlider"] > div > div > div > p {
                color: rgb(49, 51, 63) !important;  /* Cor padrão do texto do Streamlit */
            }
            
            /* Manter apenas os valores em verde */
            [data-testid="stSlider"] [data-testid="stThumbValue"] {
                color: rgb(49, 170, 77) !important;
            }
                
            [data-testid="stSlider"] p {
                color: rgb(49, 51, 63) !important;
            }
            
        </style>
    """, unsafe_allow_html=True)
    
    # Filtros interativos
    st.sidebar.header('Filtros:')
    st.sidebar.markdown("""
    <div style="background-color: #f8f9fa; padding: 10px; border-radius: 5px; margin-bottom: 15px; font-size: 0.9em;">
        <p style="margin: 0; color: #666;">Aplicados ao cálculo de "Valor Total a Receber pela Empresa"</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Função auxiliar para ordenar meses no formato MM/AAAA
    def ordenar_datas(datas):
        # Obter o mês atual para comparação
        hoje = datetime.datetime.now()
        mes_atual_dt = datetime.datetime(hoje.year, hoje.month, 1)
        
        # Filtrar apenas datas válidas (no formato MM/AAAA) e não nulas/NA
        datas_validas = []
        for d in datas:
            # Pular "A definir"
            if pd.notna(d) and d != 'A definir' and re.match(r'^\d{2}/\d{4}$', str(d)):
                # Converter para objeto datetime para comparação
                try:
                    mes, ano = map(int, str(d).split('/'))
                    data_dt = datetime.datetime(ano, mes, 1)
                    # Incluir apenas datas futuras (>= mês atual)
                    if data_dt >= mes_atual_dt:
                        datas_validas.append(d)
                except (ValueError, TypeError):
                    # Ignorar datas inválidas
                    pass
                    
        if not datas_validas:
            return ['A definir'] if 'A definir' in datas else []

        # Converter para objetos datetime para ordenação
        try:
            # Use dictionary comprehension for safety
            data_map = {d: pd.to_datetime(d, format='%m/%Y') for d in datas_validas}
            # Ordenar as datas válidas
            datas_ordenadas = sorted(datas_validas, key=lambda d: data_map[d])
        except Exception as e:
            print(f"Erro ao ordenar datas: {e}. Datas: {datas_validas}")
            datas_ordenadas = sorted(datas_validas) # Fallback to string sort

        # Adicionar 'A definir' ao INÍCIO da lista, se existir na lista original 'datas'
        original_datas_set = set(datas) # Use set for faster lookup
        if 'A definir' in original_datas_set:
            datas_ordenadas = ['A definir'] + datas_ordenadas
        return datas_ordenadas

    # --- Use df_desvio for filter options ---
    meses_disponiveis_all = []
    if 'PREVISÃO DE DATA DE RECEBIMENTO' in df_desvio.columns:
        meses_disponiveis_all = ordenar_datas(df_desvio['PREVISÃO DE DATA DE RECEBIMENTO'].unique())
    else:
        meses_disponiveis_all = ['A definir'] # Fallback

    tipos_disponiveis_all = []
    if 'TIPO' in df_desvio.columns:
        tipos_disponiveis_all = sorted(df_desvio['TIPO'].unique())

    fundacoes_disponiveis_all = []
    if 'FUNDAÇÃO' in df_desvio.columns:
        fundacoes_disponiveis_all = sorted(df_desvio['FUNDAÇÃO'].unique())

    clientes_disponiveis_all = []
    if 'CLIENTE' in df_desvio.columns:
        # Ensure 'Não identificado' is handled correctly if present
        unique_clients = df_desvio['CLIENTE'].fillna('Não identificado').astype(str).unique()
        clientes_disponiveis_all = sorted([client for client in unique_clients if client != 'Não identificado'] + (['Não identificado'] if 'Não identificado' in unique_clients else []))

    # Initialize session state if not present (remains the same)
    if 'meses' not in st.session_state: st.session_state.meses = []
    if 'tipos' not in st.session_state: st.session_state.tipos = []
    if 'fundacoes' not in st.session_state: st.session_state.fundacoes = []
    if 'clientes' not in st.session_state: st.session_state.clientes = []
    if 'min_saldo' not in st.session_state: 
        # Initialize with minimum value for filtering
        if 'SALDO A RECEBER' in data.columns and pd.api.types.is_numeric_dtype(data['SALDO A RECEBER']):
            saldo_receber_agg = data['SALDO A RECEBER'] # This is the per-project total saldo
            min_saldo_val = float(saldo_receber_agg.min()) if not saldo_receber_agg.empty else 0.0
        else:
            min_saldo_val = 0.0
        st.session_state.min_saldo = min_saldo_val
    if 'max_saldo' not in st.session_state: 
        # Initialize with maximum value for filtering
        if 'SALDO A RECEBER' in data.columns and pd.api.types.is_numeric_dtype(data['SALDO A RECEBER']):
            saldo_receber_agg = data['SALDO A RECEBER'] # This is the per-project total saldo
            max_saldo_val = float(saldo_receber_agg.max()) if not saldo_receber_agg.empty else 1000000.0
        else:
            max_saldo_val = 1000000.0
        st.session_state.max_saldo = max_saldo_val

    # --- Create a form for filters to avoid reloading on every interaction ---
    with st.sidebar.form(key="filter_form"):
        # Get filter values from form using the full lists
        meses_temp = st.multiselect('Meses (Previsão):', meses_disponiveis_all, default=st.session_state.get('meses', []))
        tipos_temp = st.multiselect('Tipos de Serviço:', tipos_disponiveis_all, default=st.session_state.get('tipos', []))
        fundacoes_temp = st.multiselect('Fundações:', fundacoes_disponiveis_all, default=st.session_state.get('fundacoes', []))
        clientes_temp = st.multiselect('Clientes:', clientes_disponiveis_all, default=st.session_state.get('clientes', []))
        
        # Custom Brazilian currency formatter for min/max inputs
        def formatar_moeda_br(valor):
            return f"R$ {valor:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
            
        # Função para validar e converter input de texto para número
        def texto_para_numero(texto):
            if not texto or texto == "R$ ":
                return None
            try:
                # Remove R$, espaços e converte vírgula para ponto
                limpo = texto.replace("R$", "").replace(".", "").replace(",", ".").strip()
                return float(limpo)
            except:
                return None
        
        st.markdown('Intervalo de Saldo total dos projetos a serem exibidos:')
        
        
        # Cria duas colunas para os inputs lado a lado
        col_min, col_max = st.columns(2)
        
        # Get current values from session state
        min_saldo_current = st.session_state.get('min_saldo', 0.0)
        max_saldo_current = st.session_state.get('max_saldo', 1000.0)
        
        with col_min:
            min_input = st.text_input(
                'Valor Mínimo:',
                value="R$ ",
                placeholder=""
            )
            
            # Garantir prefixo R$
            if not min_input.startswith("R$ "):
                min_input = "R$ " + min_input.lstrip("R$ ")
            
            # Converte o input para número
            min_saldo_temp = texto_para_numero(min_input)
            if min_saldo_temp is None:
                min_saldo_temp = min_saldo_current
            
        with col_max:
            max_input = st.text_input(
                'Valor Máximo:',
                value="R$ ",
                placeholder=""
            )
            
            # Garantir prefixo R$
            if not max_input.startswith("R$ "):
                max_input = "R$ " + max_input.lstrip("R$ ")
            
            # Converte o input para número
            max_saldo_temp = texto_para_numero(max_input)
            if max_saldo_temp is None:
                max_saldo_temp = max_saldo_current

        # Submit button for the form
        submitted = st.form_submit_button("Aplicar Filtros", use_container_width=True)
        
        # Clear filters button in the form
        clear_filters = st.form_submit_button("Limpar Filtros", use_container_width=True)
    
    # Process form submission outside the form
    if submitted:
        st.session_state.meses = meses_temp
        st.session_state.tipos = tipos_temp
        st.session_state.fundacoes = fundacoes_temp
        st.session_state.clientes = clientes_temp
        st.session_state.min_saldo = min_saldo_temp
        st.session_state.max_saldo = max_saldo_temp
    
    # Clear filters if requested
    if clear_filters:
        st.session_state.meses = []
        st.session_state.tipos = []
        st.session_state.fundacoes = []
        st.session_state.clientes = []
        
        # Reset to calculated min/max values
        if 'SALDO A RECEBER' in data.columns and pd.api.types.is_numeric_dtype(data['SALDO A RECEBER']):
            saldo_receber_agg = data['SALDO A RECEBER']
            st.session_state.min_saldo = float(saldo_receber_agg.min()) if not saldo_receber_agg.empty else 0.0
            st.session_state.max_saldo = float(saldo_receber_agg.max()) if not saldo_receber_agg.empty else 1000000.0
        else:
            st.session_state.min_saldo = 0.0
            st.session_state.max_saldo = 1000000.0
    
    # Usar os filtros armazenados na sessão para filtrar os dados
    filtered_data = data.copy()
    if st.session_state.meses:
        filtered_data = filtered_data[filtered_data['DATA'].isin(st.session_state.meses)]
    if st.session_state.tipos:
        filtered_data = filtered_data[filtered_data['TIPO'].isin(st.session_state.tipos)]
    if st.session_state.fundacoes:
        filtered_data = filtered_data[filtered_data['FUNDAÇÃO'].isin(st.session_state.fundacoes)]
    if st.session_state.clientes:
        filtered_data = filtered_data[filtered_data['CLIENTE'].isin(st.session_state.clientes)]
    
    filtered_data = filtered_data[(filtered_data['SALDO A RECEBER'] >= st.session_state.min_saldo) &
                                  (filtered_data['SALDO A RECEBER'] <= st.session_state.max_saldo)]
    
    # Add a message to guide users about the form behavior
    st.sidebar.markdown("""
    <div style="background-color: #f8f9fa; padding: 10px; border-radius: 5px; margin-bottom: 15px; font-size: 0.9em;">
        <p style="margin: 0; color: #666;">💡 Selecione todos os filtros desejados antes de clicar em "Aplicar Filtros".</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.sidebar.subheader('Resumo dos Filtros:')
    st.sidebar.write('Número de projetos:', filtered_data.shape[0])

    # --- Planilha de Contas a Receber Section ---

    # 1. Prepare data and Excel buffer regardless of button press

    # --- Helper function for new column ---
    # MODIFIED: Implement new calculation logic based on proportion of costs
    def calculate_predicted_receivable(project_key, df_all_data, selected_months, project_total_saldo_receber):
        """
        Calculates the predicted receivable amount for a project based on selected months
        using the new proportion-of-costs logic.

        Args:
            project_key (str): The unique key for the project (PROJETO_ID_KEY).
            df_all_data (pd.DataFrame): The full dataframe with all rows (df_desvio).
            selected_months (list): List of months selected in the filter (e.g., ['01/2024', '02/2024']).
            project_total_saldo_receber (float): The total 'SALDO A RECEBER' for this project (used when no months are filtered).

        Returns:
            float: The calculated predicted receivable amount based on the new logic.
        """
        # If no months are filtered, return the total SALDO A RECEBER for the project
        if not selected_months:
            return project_total_saldo_receber

        # Ensure necessary columns exist in df_all_data
        required_cols = ['PROJETO_ID_KEY', 'PREVISÃO DE DATA DE RECEBIMENTO', 'PREVISÃO DE VALOR DE RECEBIMENTO',
                         'VALOR DO CONTRATO', 'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS', 'SALDO A RECEBER']
        if not all(col in df_all_data.columns for col in required_cols):
            print(f"Warning: Required columns missing in df_all_data for predicted receivable calculation for key {project_key}. Required: {required_cols}")
            return 0

        # Filter df_all_data for the specific project
        project_rows = df_all_data[df_all_data['PROJETO_ID_KEY'] == project_key]

        if project_rows.empty:
            print(f"Warning: Project key {project_key} not found in detailed data for prediction.")
            return 0

        # Ensure key numeric columns are numeric, fillna with 0
        numeric_cols_predict = ['PREVISÃO DE VALOR DE RECEBIMENTO', 'VALOR DO CONTRATO', 'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS', 'SALDO A RECEBER']
        for col in numeric_cols_predict:
            project_rows[col] = pd.to_numeric(project_rows[col], errors='coerce').fillna(0)

        # Get project-level constants (assuming they are the same for all rows of a project)
        project_valor_contrato = project_rows['VALOR DO CONTRATO'].iloc[0] if not project_rows.empty else 0
        project_custos_incorridos = project_rows['CUSTOS INCORRIDOS'].iloc[0] if not project_rows.empty else 0
        project_outros_correlatos = project_rows['OUTROS CORRELATOS'].iloc[0] if not project_rows.empty else 0
        project_total_costs = project_custos_incorridos + project_outros_correlatos

        # Separate rows with "A definir" and rows with defined dates
        project_rows['DATA_PREVISTA'] = project_rows['PREVISÃO DE DATA DE RECEBIMENTO'].astype(str).fillna('A definir')
        rows_a_definir = project_rows[project_rows['DATA_PREVISTA'] == 'A definir']
        rows_com_data = project_rows[project_rows['DATA_PREVISTA'] != 'A definir']

        # Calculate total predicted value
        total_predicted_value = 0

        # For rows with "A definir", use SALDO A RECEBER directly (no proportion calculation)
        if not rows_a_definir.empty and 'A definir' in selected_months:
            # Get unique SALDO A RECEBER for "A definir" rows to avoid duplication
            saldo_a_definir = rows_a_definir['SALDO A RECEBER'].iloc[0]
            total_predicted_value += saldo_a_definir

        # For rows with defined dates, calculate based on proportion of costs
        if not rows_com_data.empty:
            # Filter by selected months (excluding "A definir")
            selected_months_without_a_definir = [m for m in selected_months if m != 'A definir']
            if selected_months_without_a_definir:
                monthly_filtered_rows = rows_com_data[rows_com_data['DATA_PREVISTA'].isin(selected_months_without_a_definir)]
                
                for _, line_row in monthly_filtered_rows.iterrows():
                    line_previsao_valor = line_row['PREVISÃO DE VALOR DE RECEBIMENTO']

                    if project_valor_contrato > 0:
                        proportion = line_previsao_valor / project_valor_contrato
                        predicted_value_for_line = proportion * project_total_costs
                        total_predicted_value += predicted_value_for_line

        return total_predicted_value
    # --- End Helper function ---

    # --- Filtering Logic ---
    # Start with the raw data containing all lines
    df_contas_base = df_desvio.copy()

    # Apply sidebar filters (Months, Types, Foundations, Clients)
    if st.session_state.meses:
        # Ensure filter works even if column has NaT or mixed types temporarily
        date_col_filter = 'PREVISÃO DE DATA DE RECEBIMENTO'
        if date_col_filter in df_contas_base.columns:
            df_contas_base[date_col_filter] = df_contas_base[date_col_filter].astype(str).fillna('A definir')
            df_contas_base = df_contas_base[df_contas_base[date_col_filter].isin(st.session_state.meses)]
        else:
             print(f"Warning: Column '{date_col_filter}' not found for month filtering.")
    if st.session_state.tipos:
        if 'TIPO' in df_contas_base.columns:
            df_contas_base = df_contas_base[df_contas_base['TIPO'].isin(st.session_state.tipos)]
    if st.session_state.fundacoes:
         if 'FUNDAÇÃO' in df_contas_base.columns:
            df_contas_base = df_contas_base[df_contas_base['FUNDAÇÃO'].isin(st.session_state.fundacoes)]
    if st.session_state.clientes:
         if 'CLIENTE' in df_contas_base.columns:
            df_contas_base['CLIENTE'] = df_contas_base['CLIENTE'].astype(str).fillna('Não identificado')
            df_contas_base = df_contas_base[df_contas_base['CLIENTE'].isin(st.session_state.clientes)]

    # Apply Slider Filter based on TOTAL Project Saldo
    # 1. Use 'data' (which has aggregated saldo) to find which projects pass the slider
    # Need to ensure PROJETO_ID_KEY exists and is consistent in 'data'
    data_for_slider = data.copy() # Use the aggregated data 'data' from carregar_dados_dashboard

    # Re-generate PROJETO_ID_KEY on data_for_slider for safety, using same logic as df_desvio
    key_cols_slider = []
    if 'QUANT.' in data_for_slider.columns: key_cols_slider.append('QUANT.')
    if 'CLIENTE' in data_for_slider.columns: key_cols_slider.append('CLIENTE')
    if 'VALOR DO CONTRATO' in data_for_slider.columns: key_cols_slider.append('VALOR DO CONTRATO')

    if len(key_cols_slider) >= 2:
         data_for_slider['PROJETO_ID_KEY'] = data_for_slider.apply(
             lambda row: '_'.join(str(row[col]).strip() for col in key_cols_slider),
             axis=1
         )
         # Ensure SALDO A RECEBER is numeric before filtering
         if 'SALDO A RECEBER' in data_for_slider.columns:
              data_for_slider['SALDO A RECEBER'] = pd.to_numeric(data_for_slider['SALDO A RECEBER'], errors='coerce').fillna(0)
              filtered_projects_by_slider = data_for_slider[
                  (data_for_slider['SALDO A RECEBER'] >= st.session_state.min_saldo) &
                  (data_for_slider['SALDO A RECEBER'] <= st.session_state.max_saldo)
              ]
              valid_project_keys_from_slider = filtered_projects_by_slider['PROJETO_ID_KEY'].unique()

              # 2. Filter the df_contas_base (which has all lines) using the keys from the slider filter
              if 'PROJETO_ID_KEY' in df_contas_base.columns:
                   df_contas_filtered = df_contas_base[df_contas_base['PROJETO_ID_KEY'].isin(valid_project_keys_from_slider)].copy() # Use copy to avoid SettingWithCopyWarning
              else:
                   print("Error: 'PROJETO_ID_KEY' not found in df_contas_base for slider filtering.")
                   df_contas_filtered = df_contas_base.copy() # Fallback: keep all rows if key missing
         else:
              print("Error: 'SALDO A RECEBER' not found in data_for_slider for slider filtering.")
              df_contas_filtered = df_contas_base.copy() # Fallback
    else:
         print("Error: Could not generate PROJETO_ID_KEY in data_for_slider. Slider filter skipped.")
         df_contas_filtered = df_contas_base.copy() # Fallback

    # --- Prepare final DataFrame for display (df_exibir_final_contas) ---
    df_exibir_final_contas = df_contas_filtered.copy() # Start with the fully filtered data (all lines)
    df_exibir_final_contas = replace_none_with_dash(df_exibir_final_contas) # Apply dash replacement

    # Ensure PROJETO_ID_KEY exists (it should from df_desvio)
    if 'PROJETO_ID_KEY' not in df_exibir_final_contas.columns:
         st.error("Erro crítico: 'PROJETO_ID_KEY' não encontrado nos dados filtrados para a Planilha de Contas a Receber.")
         # Handle error gracefully, maybe display an empty dataframe or stop
         df_exibir_final_contas = pd.DataFrame() # Show empty
    else:
        # Create user-friendly 'ID PROJETO' for grouping display
        group_cols_exibir_contas = ['CLIENTE', 'PROJETO', 'VALOR DO CONTRATO']
        valid_group_cols = [col for col in group_cols_exibir_contas if col in df_exibir_final_contas.columns]
        if valid_group_cols:
            # Use factorize for potentially better performance than ngroup on large data
            # Factorize based on the combination of grouping columns
            group_tuples = df_exibir_final_contas[valid_group_cols].apply(tuple, axis=1)
            df_exibir_final_contas['ID PROJETO_NUM'] = pd.factorize(group_tuples)[0] + 1
            df_exibir_final_contas['ID PROJETO'] = df_exibir_final_contas['ID PROJETO_NUM'].apply(lambda x: f"Projeto #{x}")
            df_exibir_final_contas = df_exibir_final_contas.sort_values(by='ID PROJETO_NUM') # Sort by numeric ID first
        else:
            df_exibir_final_contas['ID PROJETO'] = "Projeto #N/A" # Fallback ID
            df_exibir_final_contas = df_exibir_final_contas.sort_index() # Sort by index if grouping fails
            print("Warning: Could not group by standard columns for ID PROJETO in Contas a Receber.")

        # --- Calculate the new column AFTER filtering ---
        # We need the total saldo for each project first (for the case where no months are filtered)
        # Use 'data' again to get the correct total saldo per project key
        saldo_map = data_for_slider.set_index('PROJETO_ID_KEY')['SALDO A RECEBER'].to_dict()

        predicted_values = {}
        for idx, row in df_exibir_final_contas.iterrows():
            key = row['PROJETO_ID_KEY']
            # Get the project's total saldo from the map (derived from 'data')
            total_saldo = saldo_map.get(key, 0) # Default to 0 if key not found
            # Pass the full df_desvio for calculation context
            predicted_values[idx] = calculate_predicted_receivable(key, df_desvio, st.session_state.meses, total_saldo)

        # Assign the calculated values using the DataFrame index
        df_exibir_final_contas['SALDO_PREVISTO_FILTRADO'] = pd.Series(predicted_values)
        df_exibir_final_contas['SALDO_PREVISTO_FILTRADO'] = df_exibir_final_contas['SALDO_PREVISTO_FILTRADO'].fillna(0)

        df_exibir_final_contas = df_exibir_final_contas.rename(columns={'SALDO_PREVISTO_FILTRADO': 'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'})

        # --- Coloring logic ---
        # Use the same INTERNAL_PROJECT_ID for consistent coloring across tables
        df_exibir_final_contas['INTERNAL_PROJECT_ID'] = df_exibir_final_contas['PROJETO_ID_KEY']
        unique_projects_contas = df_exibir_final_contas['INTERNAL_PROJECT_ID'].unique()
        color_map_rgba_contas = get_global_color_mapping(unique_projects_contas, style='rgba')
        # Map INTERNAL_PROJECT_ID to color for styling function
        key_to_color_contas = {key: color_map_rgba_contas.get(key, 'rgba(255, 255, 255, 0)') for key in unique_projects_contas}

        # Define available columns for display
        colunas_ordem_contas = [
            'ID PROJETO', 'PÁGINA', 'FUNDAÇÃO', 'CLIENTE', 'PROJETO',
            'NOMENCLATURA DO PROJETO', 'TIPO', 'CONTRATO', 'Nº TED',
            'SECRETARIA',
            'PREVISÃO DE DATA DE RECEBIMENTO', 'PREVISÃO DE VALOR DE RECEBIMENTO',
            'VALOR DO CONTRATO',
            'REPASSE RECEBIDO', 'DATA DE RECEBIMENTO',
            'SALDO A RECEBER DO CONTRATO',
            'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS',
            'EMITIDO INCORRIDOS', 'EMITIDO CORRELATOS',
            'SALDO A RECEBER',
            'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'
        ]

        # Rename columns
        df_exibir_final_contas = df_exibir_final_contas.rename(columns={
            'VALOR': 'EMITIDO INCORRIDOS',
            'VALOR2': 'EMITIDO CORRELATOS'
        })

        # Filter available columns and maintain order
        colunas_disponiveis_contas = [col for col in colunas_ordem_contas if col in df_exibir_final_contas.columns]
        
        # Create display dataframe with ordered columns
        df_display_contas = df_exibir_final_contas[colunas_disponiveis_contas].copy()

        # --- Update Excel export settings ---
        hex_color_map_filtrado = get_global_color_mapping(unique_projects_contas, style='hex')
        
        # Define columns in the exact same order as the display table
        colunas_ordem_contas = [
            'ID PROJETO', 'PÁGINA', 'FUNDAÇÃO', 'CLIENTE', 'PROJETO',
            'NOMENCLATURA DO PROJETO', 'TIPO', 'CONTRATO', 'Nº TED',
            'SECRETARIA',
            'PREVISÃO DE DATA DE RECEBIMENTO', 'PREVISÃO DE VALOR DE RECEBIMENTO',
            'VALOR DO CONTRATO',
            'REPASSE RECEBIDO', 'DATA DE RECEBIMENTO',
            'SALDO A RECEBER DO CONTRATO',
            'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS',
            'EMITIDO INCORRIDOS', 'EMITIDO CORRELATOS',
            'SALDO A RECEBER',
            'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'
        ]

        # Filter available columns and maintain order
        colunas_disponiveis_contas = [col for col in colunas_ordem_contas if col in df_exibir_final_contas.columns]
        
        # Create export dataframe with ordered columns
        df_export_filtrado = df_exibir_final_contas[colunas_disponiveis_contas + ['INTERNAL_PROJECT_ID']].copy()

        # Update currency columns based on the ordered list
        currency_cols_filtrado = [
            'PREVISÃO DE VALOR DE RECEBIMENTO', 'VALOR DO CONTRATO', 'REPASSE RECEBIDO',
            'SALDO A RECEBER DO CONTRATO', 'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS',
            'EMITIDO INCORRIDOS', 'EMITIDO CORRELATOS', 'SALDO A RECEBER',
            'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'
        ]
        numeric_cols_filtrado = [] # No numeric columns that need special formatting
        percentage_cols_filtrado = [] # No percentage columns

        # Ensure project_id_col exists before passing to excel function
        project_id_col_for_excel = 'INTERNAL_PROJECT_ID' if 'INTERNAL_PROJECT_ID' in df_export_filtrado.columns else None

        excel_buffer_filtrado = create_styled_excel(
            df_export_filtrado, # Use the dataframe prepared for export
            project_id_col=project_id_col_for_excel, # Pass the correct key column name or None
            color_mapping=hex_color_map_filtrado,
            numeric_cols=numeric_cols_filtrado,
            currency_cols=[col for col in currency_cols_filtrado if col in df_export_filtrado.columns], # Filter based on actual columns
            percentage_cols=percentage_cols_filtrado,
            filename="contas_a_receber_filtrado.xlsx",
            drop_id_col_on_export=True # Drop the INTERNAL_PROJECT_ID from final export
        )

        # --- Calculate total value BEFORE displaying the table ---
        # Initial calculation based on filtered_data (covers no-filter/'A definir' cases)
        group_cols_total_initial = ['QUANT.', 'CLIENTE', 'PROJETO']
        if all(col in filtered_data.columns for col in group_cols_total_initial):
            # Group by project to avoid double counting SALDO A RECEBER when multiple rows exist per project
            df_grouped_total_initial = filtered_data.groupby(group_cols_total_initial)['SALDO A RECEBER'].first().reset_index()
            total_a_receber_display_initial = df_grouped_total_initial['SALDO A RECEBER'].sum()
        else:
            print("Warning: Columns for grouping initial total not found. Calculation might be inaccurate.")
            # Fallback: Sum directly, might overcount if project rows are duplicated in filtered_data
            total_a_receber_display_initial = filtered_data['SALDO A RECEBER'].sum() if 'SALDO A RECEBER' in filtered_data.columns else 0

        total_a_receber_formatado_initial = f'R$ {total_a_receber_display_initial:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
        st.sidebar.write('Total a Receber (Filtros Aplicados):', total_a_receber_formatado_initial) # Keep sidebar updated

        # --- Display total value section BEFORE buttons ---
        st.subheader('Valor Total a Receber pela Empresa:')
        # Use a container or placeholder to update the value later if needed
        total_value_placeholder = st.empty()
        total_value_placeholder.write(f'<p style="font-size:40px">{total_a_receber_formatado_initial}</p>', unsafe_allow_html=True)

        # --- Buttons for filtered spreadsheet ---
        # 2. Place Download button directly (uses excel_buffer_filtrado with all rows)
        st.download_button(
            label="📥 Download Planilha Filtrada (Excel)",
            data=excel_buffer_filtrado,
            file_name=f"contas_a_receber_filtrado_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # 3. Place Show button directly
        show_table_filtrado = st.button('Mostrar Planilha Filtrada', key='btn_mostrar_planilha')
        
        # Store display state in session_state to prevent refreshes when other widgets are used
        if show_table_filtrado:
            st.session_state.show_main_table = True
        
        # Check if 'show_main_table' exists in session state
        if 'show_main_table' not in st.session_state:
            st.session_state.show_main_table = False
            
        # If "Limpar" button is pressed in another section, reset this display too
        if clear_filters:
            st.session_state.show_main_table = False

        # --- Prepare DataFrame for DISPLAY ONLY (Applying the 'A definir' grouping) ---
        df_processed_for_display = pd.DataFrame() # Initialize empty
        if not df_exibir_final_contas.empty:
            if 'PREVISÃO DE DATA DE RECEBIMENTO' in df_exibir_final_contas.columns and 'INTERNAL_PROJECT_ID' in df_exibir_final_contas.columns:
                # Split data
                df_com_data = df_exibir_final_contas[df_exibir_final_contas['PREVISÃO DE DATA DE RECEBIMENTO'] != 'A definir'].copy()
                df_a_definir = df_exibir_final_contas[df_exibir_final_contas['PREVISÃO DE DATA DE RECEBIMENTO'] == 'A definir'].copy()

                # Group 'A definir' rows by project ID and take the first row
                if not df_a_definir.empty:
                    df_a_definir_grouped = df_a_definir.groupby('INTERNAL_PROJECT_ID', observed=True, dropna=False).first().reset_index()
                    # Recombine
                    df_processed_for_display = pd.concat([df_com_data, df_a_definir_grouped], ignore_index=True)
                else:
                    # If no 'A definir' rows, just use the df_com_data
                    df_processed_for_display = df_com_data

                # Re-sort based on the original sorting logic if needed (e.g., by ID PROJETO_NUM)
                if 'ID PROJETO_NUM' in df_processed_for_display.columns:
                     df_processed_for_display = df_processed_for_display.sort_values(by='ID PROJETO_NUM').reset_index(drop=True)
                else:
                     # Fallback sort if ID PROJETO_NUM was lost or not generated
                     df_processed_for_display = df_processed_for_display.sort_index()
                
                # CORREÇÃO: Recalcular o total quando houver apenas datas selecionadas (Moved to update placeholder)
                # if st.session_state.meses and all(m != 'A definir' for m in st.session_state.meses) and 'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA' in df_processed_for_display.columns:
                #     # Se tivermos apenas filtros de data (sem "A definir"), recalcular o total baseado nos dados processados
                #     total_a_receber_display = df_processed_for_display['SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'].sum()
                #     # Atualizar e exibir o valor correto
                #     total_a_receber_formatado = f'R$ {total_a_receber_display:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                #     st.subheader('Valor Total a Receber pela Empresa:') # This display is removed
                #     st.write(f'<p style="font-size:40px">{total_a_receber_formatado}</p>', unsafe_allow_html=True) # This display is removed
            else:
                 print("Warning: 'PREVISÃO DE DATA DE RECEBIMENTO' or 'INTERNAL_PROJECT_ID' missing. Displaying original filtered data.")
                 df_processed_for_display = df_exibir_final_contas.copy() # Fallback to original if columns missing
        # --- End Preparation for Display ---


        # --- Re-calculate and update total value if ONLY date filters are applied ---
        only_date_filter_applied = st.session_state.meses and all(m != 'A definir' for m in st.session_state.meses)
        if only_date_filter_applied and 'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA' in df_processed_for_display.columns:
            # Se tivermos apenas filtros de data (sem "A definir"), recalcular o total baseado nos dados processados
            total_a_receber_recalculated = df_processed_for_display['SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'].sum()
            # Atualizar e exibir o valor correto no placeholder
            total_a_receber_formatado_recalculated = f'R$ {total_a_receber_recalculated:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            total_value_placeholder.write(f'<p style="font-size:40px">{total_a_receber_formatado_recalculated}</p>', unsafe_allow_html=True)
            # Optionally update sidebar as well if needed
            st.sidebar.write('Total a Receber (Filtros Aplicados):', total_a_receber_formatado_recalculated) # Update sidebar too
        # --- End Re-calculation ---

        # --- Display filtered spreadsheet table conditionally ---
        # 4. Display table conditionally based on the "Mostrar" button
        # Use df_processed_for_display for the table content
        if st.session_state.show_main_table and not df_processed_for_display.empty: # Check if there's data to display
            st.markdown("<h3 style='font-size:140%;'>Planilha de Contas a Receber - Tratada/Higienizada</h3>", unsafe_allow_html=True)
            st.info(
                """
                *   Exibe registros de **contas a receber** com base nos filtros aplicados.
                *   Linhas pertencentes ao mesmo projeto são **agrupadas visualmente** por cor e identificadas por um **ID PROJETO** único.
                *   **Atenção:** Para projetos com previsão 'A definir', **apenas uma linha resumo** é exibida nesta tabela. A exportação para Excel contém todas as linhas originais.
                *   A coluna **'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'** mostra o valor total a receber do projeto ('SALDO A RECEBER') se nenhum mês for filtrado. Se meses forem filtrados, calcula um valor previsto baseado na proporção dos custos para as parcelas nos meses selecionados (ou usa o saldo total para 'A definir').
                *   A coluna **'SALDO A RECEBER'** representa o saldo *total* restante para o projeto.
                """
            )

            # Define columns for display (should match previous list, excluding internal IDs if not needed)
            colunas_ordem_display_contas = [
                'ID PROJETO', 'PÁGINA', 'FUNDAÇÃO', 'CLIENTE', 'PROJETO',
                'NOMENCLATURA DO PROJETO', 'TIPO', 'CONTRATO', 'Nº TED',
                'SECRETARIA',
                'PREVISÃO DE DATA DE RECEBIMENTO', 'PREVISÃO DE VALOR DE RECEBIMENTO',
                'VALOR DO CONTRATO',
                'REPASSE RECEBIDO', 'DATA DE RECEBIMENTO',
                'SALDO A RECEBER DO CONTRATO',
                'CUSTOS INCORRIDOS', 'OUTROS CORRELATOS',
                'EMITIDO INCORRIDOS', 'EMITIDO CORRELATOS',
                'SALDO A RECEBER',
                'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA'
            ]
            # Select only available columns from the processed dataframe
            colunas_disponiveis_display = [col for col in colunas_ordem_display_contas if col in df_processed_for_display.columns]
            df_display_contas = df_processed_for_display[colunas_disponiveis_display + ['INTERNAL_PROJECT_ID']].copy() # Keep ID for styling

            # --- Update formatting dictionary based on new columns/order ---
            format_dict = {
                'PREVISÃO DE VALOR DE RECEBIMENTO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'VALOR DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'REPASSE RECEBIDO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'SALDO A RECEBER DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'SALDO A RECEBER': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'CUSTOS INCORRIDOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'OUTROS CORRELATOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'EMITIDO INCORRIDOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—", # Renamed col
                'EMITIDO CORRELATOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—", # Renamed col
                'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                # Add formatting for date columns if needed, e.g., ensuring consistent display
                'PREVISÃO DE DATA DE RECEBIMENTO': lambda x: str(x) if pd.notna(x) else "A definir",
                'DATA DE RECEBIMENTO': lambda x: str(x) if pd.notna(x) else "—",
                'PÁGINA': lambda x: str(x) if pd.notna(x) else "—",
                'FUNDAÇÃO': lambda x: str(x) if pd.notna(x) else "—",
                'CLIENTE': lambda x: str(x) if pd.notna(x) else "—",
                'PROJETO': lambda x: str(x) if pd.notna(x) else "—",
                'NOMENCLATURA DO PROJETO': lambda x: str(x) if pd.notna(x) else "—",
                'TIPO': lambda x: str(x) if pd.notna(x) else "—",
                'CONTRATO': lambda x: str(x) if pd.notna(x) else "—",
                'Nº TED': lambda x: str(x) if pd.notna(x) else "—",
                'SECRETARIA': lambda x: str(x) if pd.notna(x) else "—",
                'ID PROJETO': lambda x: str(x) if pd.notna(x) else "—"
            }
            # Filter format_dict to only include columns present in df_display_contas
            valid_format_dict = {k: v for k, v in format_dict.items() if k in df_display_contas.columns}
            # --- Update formatting dictionary based on new columns/order ---
            format_dict = {
                'PREVISÃO DE VALOR DE RECEBIMENTO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'VALOR DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'REPASSE RECEBIDO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'SALDO A RECEBER DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'SALDO A RECEBER': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'CUSTOS INCORRIDOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'OUTROS CORRELATOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                'EMITIDO INCORRIDOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—", # Renamed col
                'EMITIDO CORRELATOS': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—", # Renamed col
                'SALDO A RECEBER PREVISTO ATÉ A DATA FILTRADA': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) and isinstance(x, (int, float)) else "—",
                # Add formatting for date columns if needed, e.g., ensuring consistent display
                'PREVISÃO DE DATA DE RECEBIMENTO': lambda x: str(x) if pd.notna(x) else "A definir",
                'DATA DE RECEBIMENTO': lambda x: str(x) if pd.notna(x) else "—",
                'PÁGINA': lambda x: str(x) if pd.notna(x) else "—",
                'FUNDAÇÃO': lambda x: str(x) if pd.notna(x) else "—",
                'CLIENTE': lambda x: str(x) if pd.notna(x) else "—",
                'PROJETO': lambda x: str(x) if pd.notna(x) else "—",
                'NOMENCLATURA DO PROJETO': lambda x: str(x) if pd.notna(x) else "—",
                'TIPO': lambda x: str(x) if pd.notna(x) else "—",
                'CONTRATO': lambda x: str(x) if pd.notna(x) else "—",
                'Nº TED': lambda x: str(x) if pd.notna(x) else "—",
                'SECRETARIA': lambda x: str(x) if pd.notna(x) else "—",
                'ID PROJETO': lambda x: str(x) if pd.notna(x) else "—"
            }
            # Filter format_dict to only include columns present in df_display_contas
            valid_format_dict = {k: v for k, v in format_dict.items() if k in df_display_contas.columns}

            # Get color map based on the IDs present in the processed display data
            unique_ids_display = df_display_contas['INTERNAL_PROJECT_ID'].unique()
            color_map_rgba_display = get_global_color_mapping(unique_ids_display, style='rgba')
            key_to_color_display = {key: color_map_rgba_display.get(key, 'rgba(255, 255, 255, 0)') for key in unique_ids_display}

            # Simplified styling function using the column directly
            def highlight_projects_display(row):
                key = row.get('INTERNAL_PROJECT_ID', None) # Get ID safely
                cor_rgba = key_to_color_display.get(key, 'rgba(255, 255, 255, 0)') # Default to transparent
                return [f'background-color: {cor_rgba}'] * len(row)

            # Apply styling and formatting, then drop the internal ID before display
            st.dataframe(
                df_display_contas # Use the DataFrame with INTERNAL_PROJECT_ID
                .style
                .apply(highlight_projects_display, axis=1) # Apply color based on INTERNAL_PROJECT_ID column
                .format(valid_format_dict, na_rep="—"), # Apply formatting
                column_config={ # Hide the internal ID column from view
                        "INTERNAL_PROJECT_ID": None
                },
                use_container_width=True,
                hide_index=True # Hide index for cleaner display
            )

            # Calculate and display detailed metrics based on the DISPLAYED data
            num_projetos_display = df_display_contas['ID PROJETO'].nunique()
            num_linhas_display = len(df_display_contas)
            # Metrics based on PREVISÃO DE VALOR DE RECEBIMENTO might be less relevant now for 'A definir'
            # Let's adjust the metric description or calculation if needed

            st.markdown("""
            <div style='font-size:100%;'>
                <p>📊 Resumo da Planilha Exibida:</p>
                <ul>
                    <li>Número de Projetos Agregados: {}</li>
                    <li>Número Total de Linhas Exibidas: {}</li>
                </ul>
            </div>
            """.format(num_projetos_display, num_linhas_display), unsafe_allow_html=True)
            
            # Add a button to close the table
            if st.button("Fechar Tabela", key="btn_fechar_main_table"):
                st.session_state.show_main_table = False
                st.rerun()  # Force rerun to hide the table
                    


        # --- End Planilha de Contas a Receber Section ---

        # Dashboard de Alerta para Saldos em Atraso
        st.markdown("---")
        st.subheader("Repasses em Atraso ⚠️")
        
        # Filtros em linha, abaixo do título
        col_filtro_instituto, col_filtro_tipo, col_filtro_meses = st.columns(3)
            
        with col_filtro_instituto:
            institutos_disponiveis = sorted(df_desvio['PÁGINA'].unique()) if 'PÁGINA' in df_desvio.columns else []
            institutos_selecionados = st.multiselect(
                "Instituto", 
                institutos_disponiveis, 
                default=[],
                key="filtro_instituto_atrasos"
            )
        
        with col_filtro_tipo:
            tipos_atraso_disponiveis = sorted(df_desvio['TIPO'].unique()) if 'TIPO' in df_desvio.columns else []
            tipos_atraso_selecionados = st.multiselect(
                "Tipo", 
                tipos_atraso_disponiveis, 
                default=[],
                key="filtro_tipo_atrasos"
            )
            
        with col_filtro_meses:
            opcoes_meses_atraso = ["até 1 mês", "até 3 meses", "até 6 meses", "até 9 meses", "até 12 meses", "Mais de 12 meses"]
            meses_atraso_selecionado = st.multiselect(
                "Meses em atraso",
                opcoes_meses_atraso,
                default=[],
                key="filtro_meses_atraso"
            )
        
        # Obter mês atual (primeiro dia do mês para comparação consistente)
        hoje = datetime.datetime.now()
        mes_atual = hoje.strftime('%m/%Y')
        # Garantir que estamos usando o primeiro dia do mês atual para comparação
        mes_atual_dt = datetime.datetime(hoje.year, hoje.month, 1)
        
        # Usar a mesma fonte de dados do desvio para garantir consistência
        df_atrasos = df_desvio.copy()
        
        # Aplicar filtros de Instituto e Tipo se selecionados
        if institutos_selecionados:
            df_atrasos = df_atrasos[df_atrasos['PÁGINA'].isin(institutos_selecionados)]
        if tipos_atraso_selecionados:
            df_atrasos = df_atrasos[df_atrasos['TIPO'].isin(tipos_atraso_selecionados)]
        
        # Verificar quais linhas têm data de previsão anterior ao mês atual E não têm repasse recebido
        # Verificar se as colunas necessárias existem e adicionar tratamento de erro
        if 'PREVISÃO DE DATA DE RECEBIMENTO' not in df_atrasos.columns:
            st.warning("Coluna 'PREVISÃO DE DATA DE RECEBIMENTO' não encontrada. Verifique a planilha de origem.")
            df_atrasos['PREVISÃO DE DATA DE RECEBIMENTO'] = 'A definir'  # Valor padrão
        
        # Função para converter string de data no formato MM/YYYY para datetime (primeiro dia do mês)
        def converter_para_data(data_str):
            if pd.isna(data_str) or data_str == 'A definir':
                return pd.NaT
            try:
                # Verificar se está no formato esperado MM/YYYY
                if isinstance(data_str, str) and re.match(r'^[0-9]{2}/[0-9]{4}$', data_str):
                    mes, ano = map(int, data_str.split('/'))
                    return datetime.datetime(ano, mes, 1)
                return pd.NaT  # Retornar NaT para formatos inválidos
            except (ValueError, TypeError):
                return pd.NaT
        
        # Converter datas usando a função personalizada
        df_atrasos['DATA_PREVISTA_DT'] = df_atrasos['PREVISÃO DE DATA DE RECEBIMENTO'].apply(converter_para_data)
        
        # Adicionar coluna para debug (opcional, pode ser removido em produção)
        df_atrasos['MES_DEBUG'] = df_atrasos['PREVISÃO DE DATA DE RECEBIMENTO'].astype(str)
        df_atrasos['DATA_PREVISTA_DEBUG'] = df_atrasos['DATA_PREVISTA_DT'].astype(str)
        
        # Uma linha está atrasada se: 
        # 1. A data prevista é anterior à data atual
        # 2. Não há valor em "Repasse Recebido" (ou é muito baixo) naquela linha
        # 3. A data prevista não é nula (não é "A definir")
        df_atrasos['LINHA_ATRASADA'] = (
            (df_atrasos['DATA_PREVISTA_DT'] < mes_atual_dt) & 
            ((df_atrasos['REPASSE RECEBIDO'].isna()) | (df_atrasos['REPASSE RECEBIDO'] < 1.0)) &
            (df_atrasos['DATA_PREVISTA_DT'].notna())
        )
        
        # Filtrar apenas projetos com saldo a receber maior que zero
        df_atrasos = df_atrasos[df_atrasos['SALDO A RECEBER'] > 0]
        
        # Calcular meses em atraso
        df_atrasos['MESES_ATRASO'] = df_atrasos.apply(
            lambda row: ((mes_atual_dt.year - row['DATA_PREVISTA_DT'].year) * 12 + 
                        (mes_atual_dt.month - row['DATA_PREVISTA_DT'].month))
            if pd.notna(row['DATA_PREVISTA_DT']) and row['LINHA_ATRASADA'] else 0, 
            axis=1
        )
        
        # Aplicar filtro de meses em atraso (se houver seleção)
        df_atrasos_para_identificar_projetos = df_atrasos.copy()
        if meses_atraso_selecionado:
            # Garantir que estamos olhando apenas para linhas com atraso real
            mascara_atraso_real = df_atrasos_para_identificar_projetos['MESES_ATRASO'] > 0
            
            # Construir uma máscara combinada para todas as seleções
            final_mask = pd.Series(False, index=df_atrasos_para_identificar_projetos.index)
            
            for selecionado in meses_atraso_selecionado:
                if selecionado == "até 1 mês":
                    final_mask |= (mascara_atraso_real & (df_atrasos_para_identificar_projetos['MESES_ATRASO'] <= 1))
                elif selecionado == "até 3 meses":
                    final_mask |= (mascara_atraso_real & (df_atrasos_para_identificar_projetos['MESES_ATRASO'] <= 3))
                elif selecionado == "até 6 meses":
                    final_mask |= (mascara_atraso_real & (df_atrasos_para_identificar_projetos['MESES_ATRASO'] <= 6))
                elif selecionado == "até 9 meses":
                    final_mask |= (mascara_atraso_real & (df_atrasos_para_identificar_projetos['MESES_ATRASO'] <= 9))
                elif selecionado == "até 12 meses":
                    final_mask |= (mascara_atraso_real & (df_atrasos_para_identificar_projetos['MESES_ATRASO'] <= 12))
                elif selecionado == "Mais de 12 meses":
                    final_mask |= (df_atrasos_para_identificar_projetos['MESES_ATRASO'] > 12)
            
            # Aplicar o filtro de meses
            df_atrasos_para_identificar_projetos = df_atrasos_para_identificar_projetos[final_mask]
            
            # Garantir que as colunas necessárias existam no DataFrame filtrado
            # 1. Primeiro, calcular PERC_REPASSE_PREVISTO se não existir
            if 'PERC_REPASSE_PREVISTO' not in df_atrasos_para_identificar_projetos.columns:
                df_atrasos_para_identificar_projetos['PERC_REPASSE_PREVISTO'] = df_atrasos_para_identificar_projetos.apply(
                    lambda row: row['PREVISÃO DE VALOR DE RECEBIMENTO'] / row['VALOR DO CONTRATO'] 
                    if pd.notna(row['PREVISÃO DE VALOR DE RECEBIMENTO']) and row['VALOR DO CONTRATO'] > 0 else 0,
                    axis=1
                )
            
            # 2. Depois, calcular SALDO_A_RECEBER_ATRASADO usando PERC_REPASSE_PREVISTO
            if 'SALDO_A_RECEBER_ATRASADO' not in df_atrasos_para_identificar_projetos.columns:
                df_atrasos_para_identificar_projetos['SALDO_A_RECEBER_ATRASADO'] = df_atrasos_para_identificar_projetos.apply(
                    lambda row: row['SALDO A RECEBER'] * row['PERC_REPASSE_PREVISTO'] 
                    if row['LINHA_ATRASADA'] else 0,
                    axis=1
                )

        # Verificar se a coluna de previsão de valor existe
        if 'PREVISÃO DE VALOR DE RECEBIMENTO' not in df_atrasos.columns:
            st.warning("Coluna 'PREVISÃO DE VALOR DE RECEBIMENTO' não encontrada. Verifique a planilha de origem.")
            df_atrasos['PREVISÃO DE VALOR DE RECEBIMENTO'] = df_atrasos['VALOR DO CONTRATO'] # Valor padrão
        
        # Calcular o saldo a receber atrasado proporcional para linhas atrasadas
        df_atrasos['PERC_REPASSE_PREVISTO'] = df_atrasos.apply(
            lambda row: row['PREVISÃO DE VALOR DE RECEBIMENTO'] / row['VALOR DO CONTRATO'] 
            if pd.notna(row['PREVISÃO DE VALOR DE RECEBIMENTO']) and row['VALOR DO CONTRATO'] > 0 else 0,
            axis=1
        )
        
        # Calcular o saldo a receber atrasado como a proporção do saldo total
        df_atrasos['SALDO_A_RECEBER_ATRASADO'] = df_atrasos.apply(
            lambda row: row['SALDO A RECEBER'] * row['PERC_REPASSE_PREVISTO'] 
            if row['LINHA_ATRASADA'] else 0,
            axis=1
        )
        
        # Identificar projetos que têm pelo menos uma linha atrasada
        projetos_com_atraso = df_atrasos_para_identificar_projetos[df_atrasos_para_identificar_projetos['LINHA_ATRASADA']].groupby(['QUANT.', 'CLIENTE', 'PROJETO']).size().reset_index()
        projetos_com_atraso.rename(columns={0: 'NUM_LINHAS_ATRASADAS'}, inplace=True)
        
        # Se há projetos com atraso, mostrar a tabela
        if not projetos_com_atraso.empty:
            # Vamos incluir TODAS as linhas dos projetos que têm pelo menos uma linha atrasada
            projeto_ids_com_atraso = list(zip(
                projetos_com_atraso['QUANT.'],
                projetos_com_atraso['CLIENTE'],
                projetos_com_atraso['PROJETO']
            ))

            linhas_para_mostrar = []
            if meses_atraso_selecionado:
                # Se um filtro de mês foi selecionado, a tabela mostrará apenas as linhas que correspondem ao filtro.
                # O df_atrasos_para_identificar_projetos já contém exatamente estas linhas, e apenas as atrasadas.
                linhas_para_mostrar.append(df_atrasos_para_identificar_projetos)
            else:
                # Se NENHUM filtro de mês foi selecionado, mantenha o comportamento original:
                # mostre TODAS as linhas de projetos que têm PELO MENOS UMA linha em atraso.
                projeto_ids_com_atraso_loop = list(zip(
                    projetos_com_atraso['QUANT.'],
                    projetos_com_atraso['CLIENTE'],
                    projetos_com_atraso['PROJETO']
                ))

                for projeto_id in projeto_ids_com_atraso_loop:
                    quant, cliente, projeto = projeto_id
                    projeto_mask = (
                        (df_atrasos['QUANT.'] == quant) &
                        (df_atrasos['CLIENTE'] == cliente) &
                        (df_atrasos['PROJETO'] == projeto)
                    )
                    linhas_para_mostrar.append(df_atrasos[projeto_mask])

            if linhas_para_mostrar:
                # Consolidar todas as linhas em um único DataFrame para exibição
                df_linhas_mostrar = pd.concat(linhas_para_mostrar, ignore_index=True)

                # --- Calculate Metrics based on *Actually Overdue* Lines (respeitando o filtro de meses) --- #
                df_realmente_atrasado = df_atrasos_para_identificar_projetos[df_atrasos_para_identificar_projetos['LINHA_ATRASADA'] == True]

                # Metric 1 & 2: Counts
                total_projetos = len(projetos_com_atraso)
                total_linhas_atrasadas = df_realmente_atrasado.shape[0]

                # Metric 3: Sum of overdue 'PREVISÃO DE VALOR DE RECEBIMENTO'
                total_repasses_atrasado_calc = df_realmente_atrasado['PREVISÃO DE VALOR DE RECEBIMENTO'].sum()

                # Metric 4: Sum of expected emissions corresponding to overdue receipts
                df_realmente_atrasado['PROP_ATRASO'] = df_realmente_atrasado.apply(
                    lambda row: row['PREVISÃO DE VALOR DE RECEBIMENTO'] / row['VALOR DO CONTRATO']
                    if row['VALOR DO CONTRATO'] and row['VALOR DO CONTRATO'] != 0 else 0,
                    axis=1
                )
                df_realmente_atrasado['CUSTO_TOTAL_LINHA'] = df_realmente_atrasado['CUSTOS INCORRIDOS'].fillna(0) + df_realmente_atrasado['OUTROS CORRELATOS'].fillna(0)
                df_realmente_atrasado['EMISSAO_ESPERADA_ATRASADA'] = df_realmente_atrasado['CUSTO_TOTAL_LINHA'] * df_realmente_atrasado['PROP_ATRASO']

                total_emissoes_atrasadas_calc = df_realmente_atrasado['EMISSAO_ESPERADA_ATRASADA'].sum()
                # --- End Metric Calculation --- #

                # Selecionar e renomear colunas para exibição (using df_linhas_mostrar)
                colunas_exibir = {
                    'FUNDAÇÃO': 'FUNDAÇÃO',
                    'PÁGINA': 'PÁGINA',
                    'CLIENTE': 'CLIENTE',
                    'PROJETO': 'PROJETO',
                    'NOMENCLATURA DO PROJETO': 'NOMENCLATURA DO PROJETO',
                    'TIPO': 'TIPO',
                    'OBJETO': 'OBJETO',
                    'CONTRATO': 'CONTRATO',
                    'Nº TED': 'Nº TED',
                    'SECRETARIA': 'SECRETARIA',
                    'VALOR DO CONTRATO': 'VALOR DO CONTRATO',
                    'SALDO A RECEBER DO CONTRATO': 'REPASSE PENDENTE DO CONTRATO',
                    'SALDO_A_RECEBER_ATRASADO': 'SALDO A RECEBER ATRASADO',
                    'MESES_ATRASO': 'MESES EM ATRASO',
                    'PREVISÃO DE DATA DE RECEBIMENTO': 'PREVISÃO DE DATAS DE RECEBIMENTO',
                    'PREVISÃO DE VALOR DE RECEBIMENTO': 'PREVISÃO DE VALORES DE RECEBIMENTO'
                }
                
                # Filtrar apenas as colunas que existem no DataFrame
                colunas_disponiveis = [col for col in colunas_exibir.keys() if col in df_linhas_mostrar.columns]
                df_exibir = df_linhas_mostrar[colunas_disponiveis].rename(
                    columns={col: colunas_exibir[col] for col in colunas_disponiveis}
                )
                
                # Substituir valores None ou NaN por "-" em todas as colunas de texto
                colunas_texto = ['NOMENCLATURA DO PROJETO', 'Nº TED', 'SECRETARIA', 'PREVISÃO DE DATAS DE RECEBIMENTO', 'DATA DE RECEBIMENTO']
                for col in colunas_texto:
                    if col in df_exibir.columns:
                        df_exibir[col] = df_exibir[col].fillna("—")
                        # Converter valores vazios para "-"
                        df_exibir[col] = df_exibir[col].replace(r'^\s*$', "—", regex=True)
                
                # Calcular métricas totais para exibição
                total_projetos = len(projetos_com_atraso)
                total_linhas_atrasadas = projetos_com_atraso['NUM_LINHAS_ATRASADAS'].sum()
                
                # Calcular total_saldo_atrasado de forma segura, verificando se a coluna existe
                if 'LINHA_ATRASADA' in df_linhas_mostrar.columns and 'SALDO_A_RECEBER_ATRASADO' in df_linhas_mostrar.columns:
                    total_saldo_atrasado = df_linhas_mostrar[df_linhas_mostrar['LINHA_ATRASADA']]['SALDO_A_RECEBER_ATRASADO'].sum()
                else:
                    # Fallback: usar o total de repasses atrasados como aproximação
                    total_saldo_atrasado = total_repasses_atrasado_calc

                # Mostrar somente se houver saldo atrasado
                if total_saldo_atrasado > 0:
                    # Exibir métricas
                    col1, col2, col3, col4 = st.columns(4) # Changed to 4 columns
                    with col1:
                        st.metric(
                            label="Projetos com atrasos",
                            value=total_projetos
                        )
                    with col2:
                        st.metric(
                            label="Linhas de repasse atrasadas",
                            value=total_linhas_atrasadas
                        )
                    with col3:
                        st.metric(
                            label="Valor total de repasses atrasados", # Renamed label
                            value=f"R$ {total_repasses_atrasado_calc:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.') # Use new calculation
                        )
                    with col4: # Added 4th column for the new metric
                        st.metric(
                            label="Valor total de emissões atrasadas", # New metric label
                            value=f"R$ {total_emissoes_atrasadas_calc:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.') # Use new calculation
                        )

                    # --- Prepare Data for Atrasos Display and Export (using df_linhas_mostrar) --- #
                    # Ensure INTERNAL_PROJECT_ID exists for consistent coloring *before* renaming/selecting for display
                    df_linhas_mostrar = df_linhas_mostrar.reset_index(drop=True) # Ensure index is clean
                    if 'QUANT.' in df_linhas_mostrar.columns and 'CLIENTE' in df_linhas_mostrar.columns and 'PROJETO' in df_linhas_mostrar.columns:
                         df_linhas_mostrar['INTERNAL_PROJECT_ID'] = df_linhas_mostrar.apply(
                             lambda row: f"{row['QUANT.']}_{row['CLIENTE']}_{row['PROJETO']}", axis=1
                         )
                    elif 'CLIENTE' in df_linhas_mostrar.columns and 'PROJETO' in df_linhas_mostrar.columns:
                         # Fallback if QUANT. is missing
                         df_linhas_mostrar['INTERNAL_PROJECT_ID'] = df_linhas_mostrar.apply(
                             lambda row: f"{row['CLIENTE']}_{row['PROJETO']}", axis=1
                         )
                    else:
                        # If critical grouping columns are missing, create a dummy ID
                        print("Warning: Critical columns (CLIENTE, PROJETO) missing for INTERNAL_PROJECT_ID generation (Atrasos).")
                        df_linhas_mostrar['INTERNAL_PROJECT_ID'] = df_linhas_mostrar.index.astype(str)


                    # Prepare df_exibir for display (selecting and renaming columns)
                    colunas_disponiveis = [col for col in colunas_exibir.keys() if col in df_linhas_mostrar.columns]
                    df_exibir = df_linhas_mostrar[colunas_disponiveis + ['INTERNAL_PROJECT_ID']].rename( # Keep ID temporarily
                        columns={col: colunas_exibir[col] for col in colunas_disponiveis}
                    )

                    # Add user-friendly ID PROJETO for display
                    if 'VALOR DO CONTRATO' in df_exibir.columns and 'CLIENTE' in df_exibir.columns and 'PROJETO' in df_exibir.columns:
                        df_exibir['ID PROJETO'] = df_exibir.groupby(['CLIENTE', 'PROJETO', 'VALOR DO CONTRATO']).ngroup() + 1
                        df_exibir['ID PROJETO'] = df_exibir['ID PROJETO'].apply(lambda x: f"Projeto #{x}")
                    else: # Fallback if grouping columns aren't available after rename (should not happen ideally)
                         df_exibir['ID PROJETO'] = "Projeto #N/A"


                    # Sort for display
                    sort_columns_display = ['CLIENTE', 'PROJETO']
                    if 'QUANT.' in df_exibir.columns: # Check original column name before rename
                        sort_columns_display.append('QUANT.')
                    if 'MESES EM ATRASO' in df_exibir.columns:
                        sort_columns_display.append('MESES EM ATRASO')

                    sort_ascending_display = [True, True]
                    if 'QUANT.' in df_exibir.columns:
                        sort_ascending_display.append(True)
                    if 'MESES EM ATRASO' in df_exibir.columns:
                        sort_ascending_display.append(False)

                    # Check if sort columns exist in df_exibir *after* renaming
                    valid_sort_columns = [col for col in sort_columns_display if col in df_exibir.columns]
                    valid_sort_ascending = [asc for col, asc in zip(sort_columns_display, sort_ascending_display) if col in df_exibir.columns]

                    if valid_sort_columns:
                        df_exibir = df_exibir.sort_values(by=valid_sort_columns, ascending=valid_sort_ascending)


                    # Reorder columns for display
                    colunas_ordem_display = ['ID PROJETO', 'PÁGINA'] + [col for col in df_exibir.columns if col not in ['ID PROJETO', 'PÁGINA', 'INTERNAL_PROJECT_ID']]
                    df_exibir = df_exibir[colunas_ordem_display + ['INTERNAL_PROJECT_ID']] # Keep ID temporarily
                    df_exibir = df_exibir.reset_index(drop=True)


                    # --- Generate Excel Buffer for Atrasos (Moved Up & Corrected) --- #
                    # Use the df_exibir which has the correct columns and the INTERNAL_PROJECT_ID
                    df_export_atrasos = df_exibir.copy()

                    # Create hex color mapping using the consistent INTERNAL_PROJECT_ID
                    unique_projects_hex = df_export_atrasos['INTERNAL_PROJECT_ID'].unique()
                    hex_color_map_atrasos = get_global_color_mapping(unique_projects_hex, style='hex')

                    # Define column formatting based on the *renamed* columns in df_export_atrasos
                    currency_cols_atrasos = ['VALOR DO CONTRATO', 'REPASSE PENDENTE DO CONTRATO', 'SALDO A RECEBER ATRASADO', 'PREVISÃO DE VALORES DE RECEBIMENTO']
                    numeric_cols_atrasos = ['MESES EM ATRASO']
                    percentage_cols_atrasos = []


                    excel_buffer_atrasos = create_styled_excel(
                        df_export_atrasos, # Pass the df with renamed cols + internal ID
                        project_id_col='INTERNAL_PROJECT_ID', # Specify the internal ID for coloring
                        color_mapping=hex_color_map_atrasos,
                        numeric_cols=numeric_cols_atrasos,
                        currency_cols=currency_cols_atrasos,
                        percentage_cols=percentage_cols_atrasos,
                        filename="repasses_em_atraso.xlsx",
                        drop_id_col_on_export=True # Ensure the internal ID is dropped from final Excel
                    )

                    st.download_button(
                        label="📥 Download Repasses em Atraso (Excel)",
                        data=excel_buffer_atrasos,
                        file_name=f"repasses_em_atraso_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key='download_atrasos'
                    )
                    # --- End Excel Buffer --- #

                    # Button to show the table
                    show_atrasos_button = st.button('Mostrar Registros com Repasses em Atraso', key='btn_mostrar_atrasos')

                    # Store display state in session_state to prevent refreshes when other widgets are used
                    if show_atrasos_button:
                        st.session_state.show_atrasos_table = True

                    # Check if 'show_atrasos_table' exists in session state
                    if 'show_atrasos_table' not in st.session_state:
                        st.session_state.show_atrasos_table = False

                    # If "Limpar" button is pressed, reset this display too
                    if clear_filters:
                        st.session_state.show_atrasos_table = False

                    # Display table conditionally
                    if st.session_state.show_atrasos_table:
                        st.markdown("<h3 style='font-size:140%;'>Planilha de Repasses em Atraso</h3>", unsafe_allow_html=True)

                        st.info(
                            """
                            *   Identifica projetos com **pelo menos uma parcela de repasse em atraso** (data de previsão anterior ao mês atual e sem registro de recebimento).
                            *   **Todas as parcelas** do projeto são exibidas para contexto, mesmo as não atrasadas.
                            *   O 'Saldo a Receber Atrasado' é calculado **proporcionalmente** apenas para as parcelas em atraso.
                            *   Projetos são agrupados visualmente por cor e **ID PROJETO**.
                            """
                        )

                        # Generate RGBA color map for display styling
                        # Use the same INTERNAL_PROJECT_ID from df_exibir
                        unique_projects_rgba = df_exibir['INTERNAL_PROJECT_ID'].unique()
                        color_map_rgba = get_global_color_mapping(unique_projects_rgba, style='rgba')

                        # Create index-to-color mapping for the display DataFrame
                        index_to_color = {}
                        for idx, row in df_exibir.iterrows():
                             # Use get with a default for safety
                             index_to_color[idx] = color_map_rgba.get(row['INTERNAL_PROJECT_ID'], 'rgba(255, 255, 255, 0)')


                        # Define highlight function using the index mapping
                        def highlight_projects_atraso(row):
                            idx = row.name
                            cor_rgba = index_to_color.get(idx, 'rgba(255, 255, 255, 0)')
                            return [f'background-color: {cor_rgba}'] * len(row)


                        # Clean up display dataframe (remove internal ID, handle NaNs)
                        df_display = replace_none_with_dash(df_exibir.drop(columns=['INTERNAL_PROJECT_ID']))

                        # --- Re-add definitions for formatting functions --- #
                        def format_valor_atraso(x):
                            if isinstance(x, (int, float)) and x > 0:
                                return f"R$ {x:_.2f}".replace('.', ',').replace('_', '.')
                            else:
                                return "—"

                        def format_meses_atraso(x):
                            if isinstance(x, (int, float)) and x > 0:
                                return f"{int(x)} {'mês' if int(x) == 1 else 'meses'}"
                            else:
                                return "—"
                        # --- End re-added definitions --- #

                        st.dataframe(
                            df_display.style
                            .apply(highlight_projects_atraso, axis=1) # Use the correct highlight function
                            .format({
                                'VALOR DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.'),
                                'REPASSE PENDENTE DO CONTRATO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.'),
                                'SALDO A RECEBER ATRASADO': format_valor_atraso,
                                'MESES EM ATRASO': format_meses_atraso,
                                'PREVISÃO DE VALORES DE RECEBIMENTO': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.') if pd.notna(x) else "—"
                            }),
                            use_container_width=True
                        )

                        # Add a button to close the table
                        if st.button("Fechar Tabela", key="btn_fechar_atrasos_table"):
                            st.session_state.show_atrasos_table = False
                            st.rerun()  # Force rerun to hide the table

                        # --- Removed Download Button Logic from here --- #

                else:
                    st.success("Não há projetos com saldo a receber em atraso!")
                    
        else:
            st.success("Não há projetos com linhas de repasse em atraso!")
        

    

        # Dashboard de Desvio de Proporção
        st.markdown("---")
        st.subheader("Desvio na Proporção dos Repasses 🔍")

        try:
            # Calcular métricas de desvio
            tolerance = 20.0
            
            # Verificar se df_desvio possui as colunas necessárias
            colunas_necessarias_desvio = ['REPASSE RECEBIDO', 'VALOR DO CONTRATO', 'CUSTOS INCORRIDOS', 'VALOR', 'OUTROS CORRELATOS', 'VALOR2']
            colunas_faltantes = [col for col in colunas_necessarias_desvio if col not in df_desvio.columns]
            
            if colunas_faltantes:
                st.warning(f"Atenção: As seguintes colunas estão faltando nos dados: {', '.join(colunas_faltantes)}")
                st.info("Isso pode afetar os cálculos de desvio de proporção. Verifique a planilha de origem.")
                # Adicionar colunas faltantes com valores zerados para evitar erros
                for col in colunas_faltantes:
                    df_desvio[col] = 0
                    
            # Continuar com os cálculos de desvio
            df_desvio['PROP REPASSE'] = df_desvio.apply(
                lambda row: row['REPASSE RECEBIDO'] / row['VALOR DO CONTRATO']
                if row['REPASSE RECEBIDO'] and row['VALOR DO CONTRATO'] != 0 else np.nan, axis=1)

            df_desvio['EXPECTED VALOR'] = df_desvio.apply(
                lambda row: row['CUSTOS INCORRIDOS'] * row['PROP REPASSE']
                if pd.notna(row['PROP REPASSE']) and row['CUSTOS INCORRIDOS'] > 0 else np.nan, axis=1)

            df_desvio['EXPECTED VALOR2'] = df_desvio.apply(
                lambda row: row['OUTROS CORRELATOS'] * row['PROP REPASSE']
                if pd.notna(row['PROP REPASSE']) and row['OUTROS CORRELATOS'] > 0 else np.nan, axis=1)

            df_desvio['EXPECTED VALOR RND'] = df_desvio['EXPECTED VALOR'].round(2)
            df_desvio['EXPECTED VALOR2 RND'] = df_desvio['EXPECTED VALOR2'].round(2)
            df_desvio['VALOR RND'] = df_desvio['VALOR'].round(2)
            df_desvio['VALOR2 RND'] = df_desvio['VALOR2'].round(2)

            df_desvio['DESVIO VALOR'] = df_desvio.apply(
                lambda row: True if (pd.notna(row['EXPECTED VALOR RND']) and 
                                  (row['EXPECTED VALOR RND'] - row['VALOR RND'] > tolerance))
                            else False, axis=1)

            df_desvio['DESVIO VALOR2'] = df_desvio.apply(
                lambda row: True if (pd.notna(row['EXPECTED VALOR2 RND']) and 
                                  (row['EXPECTED VALOR2 RND'] - row['VALOR2 RND'] > tolerance))
                            else False, axis=1)

            df_desvio['DESVIO PROPORCAO'] = df_desvio['DESVIO VALOR'] | df_desvio['DESVIO VALOR2']

            # Calcular o desvio em reais somando as diferenças dos dois tipos
            df_desvio['DESVIO VALOR REAIS'] = df_desvio.apply(
                lambda row: (row['EXPECTED VALOR RND'] - row['VALOR RND']) 
                if pd.notna(row['EXPECTED VALOR RND']) and pd.notna(row['VALOR RND']) else 0, axis=1
            )

            df_desvio['DESVIO VALOR2 REAIS'] = df_desvio.apply(
                lambda row: (row['EXPECTED VALOR2 RND'] - row['VALOR2 RND']) 
                if pd.notna(row['EXPECTED VALOR2 RND']) and pd.notna(row['VALOR2 RND']) else 0, axis=1
            )

            # Somar os dois tipos de desvio para obter o desvio total em reais
            df_desvio['DESVIO EM REAIS'] = df_desvio['DESVIO VALOR REAIS'] + df_desvio['DESVIO VALOR2 REAIS']

            # Garantir que registros com valor total do desvio igual a zero reais não entrem no modelo
            df_desvio.loc[df_desvio['DESVIO EM REAIS'] <= 0, 'DESVIO PROPORCAO'] = False

            # Nova lógica: verificar desvios por bloco de projeto
            # Este cálculo corrige casos onde múltiplas linhas do mesmo projeto podem parecer em desvio
            # individualmente, mas quando analisadas em conjunto (como um bloco) estão corretas.
            # Exemplo: Um projeto com dois repasses (33% e 34%) pode ter múltiplos pagamentos distribuídos
            # que não correspondem linha a linha, mas que no total representam a proporção correta.

            # Agrupar por projeto para somar valores de múltiplas linhas do mesmo projeto
            group_cols = []
            if 'QUANT.' in df_desvio.columns:
                group_cols.append('QUANT.')
            if 'CLIENTE' in df_desvio.columns:
                group_cols.append('CLIENTE')
            else:
                # Adicionar log para debugging da coluna CLIENTE
                st.warning(f"DEBUG: A coluna 'CLIENTE' não foi encontrada no DataFrame df_desvio. Colunas disponíveis: {list(df_desvio.columns)}")
                # Verificar se tem alguma coluna similar
                colunas_similares = [col for col in df_desvio.columns if 'CLIENT' in col.upper()]
                if colunas_similares:
                    st.info(f"DEBUG: Colunas similares encontradas: {colunas_similares}")
                    # Usar a primeira coluna similar como fallback
                    if len(colunas_similares) > 0:
                        st.info(f"DEBUG: Usando '{colunas_similares[0]}' como fallback para 'CLIENTE'")
                        df_desvio['CLIENTE'] = df_desvio[colunas_similares[0]]
                        group_cols.append('CLIENTE')
                        
            if 'PROJETO' in df_desvio.columns:
                group_cols.append('PROJETO')
                
            # Lista de colunas para agregar
            agg_dict = {}
            if 'REPASSE RECEBIDO' in df_desvio.columns:
                agg_dict['REPASSE RECEBIDO'] = 'sum'
            if 'VALOR DO CONTRATO' in df_desvio.columns:
                agg_dict['VALOR DO CONTRATO'] = 'first'  # Assume que é o mesmo para todas as linhas do projeto
            if 'CUSTOS INCORRIDOS' in df_desvio.columns:
                agg_dict['CUSTOS INCORRIDOS'] = 'first'  # Pega apenas o primeiro valor, não soma, pois é o mesmo para todas as linhas
            if 'VALOR' in df_desvio.columns:
                agg_dict['VALOR'] = 'sum'  # Soma todos os valores recebidos para este projeto
            if 'OUTROS CORRELATOS' in df_desvio.columns:
                agg_dict['OUTROS CORRELATOS'] = 'first'  # Pega apenas o primeiro valor, não soma, pois é o mesmo para todas as linhas
            if 'VALOR2' in df_desvio.columns:
                agg_dict['VALOR2'] = 'sum'  # Soma todos os valores correlatos recebidos
            
            # Identificar coluna de ID do projeto - pode ser PROJECT ID ou PROJECT_ID
            project_id_column = None
            if 'PROJECT ID' in df_desvio.columns:
                project_id_column = 'PROJECT ID'
                agg_dict[project_id_column] = 'first'  # Identificador único do projeto
            elif 'PROJECT_ID' in df_desvio.columns:
                project_id_column = 'PROJECT_ID'
                agg_dict[project_id_column] = 'first'  # Identificador único do projeto
                
            if 'PÁGINA' in df_desvio.columns:
                agg_dict['PÁGINA'] = 'first'  # Página do projeto (usamos a primeira ocorrência)
                
            # Se não temos colunas para agrupar ou agregar, cria um dataframe vazio
            if not group_cols or not agg_dict:
                projeto_totals = pd.DataFrame()
            else:
                projeto_totals = df_desvio.groupby(group_cols).agg(agg_dict).reset_index()
                
                # Após a criação do dataframe projeto_totals, verificar se tem a coluna CLIENTE
                # Se não tiver, criar essa coluna com um valor padrão para evitar erros
                if 'CLIENTE' not in projeto_totals.columns and 'CLIENTE' in df_desvio.columns:
                    projeto_totals['CLIENTE'] = df_desvio['CLIENTE'].iloc[0] if len(df_desvio) > 0 else "Cliente não identificado"
                elif 'CLIENTE' not in projeto_totals.columns:
                    projeto_totals['CLIENTE'] = "Cliente não identificado"

            # Recalcular a proporção e desvios a nível de projeto agrupado
            projeto_totals['PROP REPASSE GRUPO'] = projeto_totals.apply(
                lambda row: row['REPASSE RECEBIDO'] / row['VALOR DO CONTRATO'] 
                if row['VALOR DO CONTRATO'] > 0 else 0, axis=1
            )

            # Calcular o valor esperado para NF de custos incorridos
            projeto_totals['EXPECTED VALOR GRUPO'] = projeto_totals.apply(
                lambda row: row['CUSTOS INCORRIDOS'] * row['PROP REPASSE GRUPO'] 
                if row['PROP REPASSE GRUPO'] > 0 else 0, axis=1
            )

            # Calcular o valor esperado para NF de custos correlatos 
            projeto_totals['EXPECTED VALOR2 GRUPO'] = projeto_totals.apply(
                lambda row: row['OUTROS CORRELATOS'] * row['PROP REPASSE GRUPO'] 
                if row['PROP REPASSE GRUPO'] > 0 else 0, axis=1
            )

            # Arredondar para comparação consistente
            projeto_totals['EXPECTED VALOR RND GRUPO'] = projeto_totals['EXPECTED VALOR GRUPO'].round(2)
            projeto_totals['EXPECTED VALOR2 RND GRUPO'] = projeto_totals['EXPECTED VALOR2 GRUPO'].round(2)
            projeto_totals['VALOR RND GRUPO'] = projeto_totals['VALOR'].round(2)
            projeto_totals['VALOR2 RND GRUPO'] = projeto_totals['VALOR2'].round(2)

            # Calcular desvios a nível de grupo
            projeto_totals['DESVIO VALOR GRUPO'] = projeto_totals.apply(
                lambda row: True if (pd.notna(row['EXPECTED VALOR RND GRUPO']) and 
                                  (row['EXPECTED VALOR RND GRUPO'] - row['VALOR RND GRUPO'] > tolerance))
                            else False, axis=1
            )

            projeto_totals['DESVIO VALOR2 GRUPO'] = projeto_totals.apply(
                lambda row: True if (pd.notna(row['EXPECTED VALOR2 RND GRUPO']) and 
                                  (row['EXPECTED VALOR2 RND GRUPO'] - row['VALOR2 RND GRUPO'] > tolerance))
                            else False, axis=1
            )

            # Flag de desvio a nível de grupo
            projeto_totals['DESVIO PROPORCAO GRUPO'] = projeto_totals['DESVIO VALOR GRUPO'] | projeto_totals['DESVIO VALOR2 GRUPO']

            # Corrigir o desvio no dataframe original com base na análise por grupo
            # Se um projeto não tem desvio no nível agrupado, corrigir o flag de desvio para todas as suas linhas
            for idx, row in projeto_totals.iterrows():
                if not row['DESVIO PROPORCAO GRUPO']:
                    # Selecionar todas as linhas deste projeto no dataframe original
                    projeto_mask = pd.Series(True, index=df_desvio.index)
                    
                    # Aplicar filtros apenas para colunas que existem
                    if 'QUANT.' in df_desvio.columns and 'QUANT.' in row:
                        projeto_mask = projeto_mask & (df_desvio['QUANT.'] == row['QUANT.'])
                    
                    if 'CLIENTE' in df_desvio.columns and 'CLIENTE' in row:
                        projeto_mask = projeto_mask & (df_desvio['CLIENTE'] == row['CLIENTE'])
                    
                    if 'PROJETO' in df_desvio.columns and 'PROJETO' in row:
                        projeto_mask = projeto_mask & (df_desvio['PROJETO'] == row['PROJETO'])
                    
                    # Desativar flag de desvio para todas as linhas deste projeto
                    df_desvio.loc[projeto_mask, 'DESVIO PROPORCAO'] = False
                    
                    # Também zerar o valor do desvio monetário para essas linhas
                    if 'DESVIO VALOR REAIS' in df_desvio.columns:
                        df_desvio.loc[projeto_mask, 'DESVIO VALOR REAIS'] = 0
                    if 'DESVIO VALOR2 REAIS' in df_desvio.columns:
                        df_desvio.loc[projeto_mask, 'DESVIO VALOR2 REAIS'] = 0
                    if 'DESVIO EM REAIS' in df_desvio.columns:
                        df_desvio.loc[projeto_mask, 'DESVIO EM REAIS'] = 0

            # Exibir métricas do dashboard de desvio
            total_registros = len(projeto_totals)
            projetos_com_desvio = projeto_totals['DESVIO PROPORCAO GRUPO'].sum()
            percentual_conformidade = ((total_registros - projetos_com_desvio) / total_registros) * 100 if total_registros > 0 else 0

            # Calcular desvio total somando os desvios de todos os projetos
            total_desvio = 0
            # desvio_medio = 0 # Removed calculation for average deviation

            for idx, row in projeto_totals.iterrows():
                if row['DESVIO PROPORCAO GRUPO']:
                    desvio_incorridos = (row['EXPECTED VALOR RND GRUPO'] - row['VALOR RND GRUPO']) if row['EXPECTED VALOR RND GRUPO'] > row['VALOR RND GRUPO'] else 0
                    desvio_correlatos = (row['EXPECTED VALOR2 RND GRUPO'] - row['VALOR2 RND GRUPO']) if row['EXPECTED VALOR2 RND GRUPO'] > row['VALOR2 RND GRUPO'] else 0
                    total_desvio += (desvio_incorridos + desvio_correlatos)

            # if projetos_com_desvio > 0: # Removed calculation for average deviation
            #     desvio_medio = total_desvio / projetos_com_desvio

            # Criando layout com 4 colunas para melhor distribuição visual (changed from 5)
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                st.metric(label="Total de Projetos", value=total_registros)
            with col2:
                st.metric(label="Projetos com Desvio", value=int(projetos_com_desvio))
            with col3:
                st.metric(label="Conformidade", value=f"{percentual_conformidade:.1f}%")
            with col4:
                st.metric(label="Desvio Total", value=f"R$ {total_desvio:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.'))
            # Removed the 5th column and metric for Desvio Médio por Projeto
            # with col5:
            #     valor_formatado = f"R$ {desvio_medio:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.') if desvio_medio > 0 else "R$ 0,00"
            #     st.metric(label="Desvio Médio por Projeto", value=valor_formatado)


            # --- Desvio de Proporção Section ---

            # 1. Prepare data and Excel buffer outside button check
            df_exibir_desvio = pd.DataFrame() # Initialize empty df for display
            df_export_desvio = pd.DataFrame() # Initialize empty df for export
            excel_buffer_desvio = BytesIO() # Initialize empty buffer
            show_desvio_buttons_and_table = False # Flag to control display of buttons and table

            # Check if there are any projects with deviation *before* attempting detailed processing
            if 'CLIENTE' in projeto_totals.columns and not projeto_totals[projeto_totals['DESVIO PROPORCAO GRUPO'] == True].empty:
                # Initial filter for projects flagged with deviation
                projetos_com_desvio_calc = projeto_totals[projeto_totals['DESVIO PROPORCAO GRUPO'] == True].copy()

                # Calculate actual monetary deviation
                projetos_com_desvio_calc['DESVIO INCORRIDOS GRUPO'] = projetos_com_desvio_calc['EXPECTED VALOR RND GRUPO'] - projetos_com_desvio_calc['VALOR RND GRUPO']
                projetos_com_desvio_calc['DESVIO CORRELATOS GRUPO'] = projetos_com_desvio_calc['EXPECTED VALOR2 RND GRUPO'] - projetos_com_desvio_calc['VALOR2 RND GRUPO']
                projetos_com_desvio_calc['DESVIO TOTAL GRUPO'] = projetos_com_desvio_calc['DESVIO INCORRIDOS GRUPO'] + projetos_com_desvio_calc['DESVIO CORRELATOS GRUPO']

                # Filter again for only those with positive monetary deviation
                projetos_com_desvio_final = projetos_com_desvio_calc[projetos_com_desvio_calc['DESVIO TOTAL GRUPO'] > 0]

                if not projetos_com_desvio_final.empty:
                    show_desvio_buttons_and_table = True # OK to show buttons/table

                    # Prepare the final DataFrame for display and export
                    # Check necessary columns exist
                    colunas_necessarias_desvio_display = ['QUANT.', 'CLIENTE', 'PROJETO', 'VALOR DO CONTRATO', 'PÁGINA', 'REPASSE RECEBIDO', 'PROP REPASSE GRUPO', 'CUSTOS INCORRIDOS', 'VALOR RND GRUPO', 'OUTROS CORRELATOS', 'VALOR2 RND GRUPO']
                    for col in colunas_necessarias_desvio_display:
                         if col not in projetos_com_desvio_final.columns:
                             projetos_com_desvio_final[col] = "—" # Add missing cols with placeholder

                    # Create user-friendly ID
                    group_cols_desvio = [col for col in ['CLIENTE', 'PROJETO', 'VALOR DO CONTRATO'] if col in projetos_com_desvio_final.columns]
                    if group_cols_desvio:
                         projetos_com_desvio_final['ID DO PROJETO'] = projetos_com_desvio_final.groupby(group_cols_desvio, observed=True, dropna=False).ngroup() + 1
                         projetos_com_desvio_final['ID DO PROJETO'] = projetos_com_desvio_final['ID DO PROJETO'].apply(lambda x: f"Projeto #{x}")
                    else:
                         projetos_com_desvio_final['ID DO PROJETO'] = "Projeto #N/A"


                    projetos_com_desvio_final = projetos_com_desvio_final.sort_values(by='ID DO PROJETO')

                    # Calculate percentages
                    projetos_com_desvio_final['PERC INCORRIDOS'] = projetos_com_desvio_final.apply(lambda row: (row['VALOR RND GRUPO'] / row['CUSTOS INCORRIDOS']) * 100 if isinstance(row.get('CUSTOS INCORRIDOS'), (int, float)) and row['CUSTOS INCORRIDOS'] > 0.01 and isinstance(row.get('VALOR RND GRUPO'), (int, float)) and row['VALOR RND GRUPO'] > 0 else float('nan'), axis=1)
                    projetos_com_desvio_final['PERC CORRELATOS'] = projetos_com_desvio_final.apply(lambda row: (row['VALOR2 RND GRUPO'] / row['OUTROS CORRELATOS']) * 100 if isinstance(row.get('OUTROS CORRELATOS'), (int, float)) and row['OUTROS CORRELATOS'] > 0.01 and isinstance(row.get('VALOR2 RND GRUPO'), (int, float)) and row['VALOR2 RND GRUPO'] > 0 else float('nan'), axis=1)


                    # Select and rename columns for display/export
                    colunas_exibir_desvio = {
                        'ID DO PROJETO': 'ID DO PROJETO', 'PÁGINA': 'Página', 'QUANT.': 'Quant.',
                        'CLIENTE': 'Cliente', 'PROJETO': 'Projeto', 'VALOR DO CONTRATO': 'Valor do Contrato',
                        'REPASSE RECEBIDO': 'Repasse Recebido', 'PROP REPASSE GRUPO': '% do Projeto',
                        'PERC INCORRIDOS': '% Repasse Incorridos', 'PERC CORRELATOS': '% Repasse Correlatos',
                        'DESVIO TOTAL GRUPO': 'Desvio Total (R$)'
                    }
                    # Ensure all keys exist before selection
                    final_display_cols = [col for col in colunas_exibir_desvio.keys() if col in projetos_com_desvio_final.columns]
                    df_exibir_desvio = projetos_com_desvio_final[final_display_cols].rename(columns=colunas_exibir_desvio)


                    # Clean up text columns
                    colunas_texto_desvio = ['Cliente', 'Projeto', 'Página', 'ID DO PROJETO']
                    for col in colunas_texto_desvio:
                        if col in df_exibir_desvio.columns:
                            df_exibir_desvio[col] = df_exibir_desvio[col].fillna("—")

                    df_exibir_desvio = df_exibir_desvio.reset_index(drop=True)
                    df_exibir_desvio.fillna(0, inplace=True) # Fill numeric NaNs resulting from calculations

                    # Prepare export DF (it's the same as display for this section)
                    df_export_desvio = df_exibir_desvio.copy()

                    # Generate Excel buffer
                    unique_project_ids_desvio_hex = df_export_desvio['ID DO PROJETO'].unique()
                    id_column_desvio_hex = 'ID DO PROJETO'

                    hex_color_map_desvio = get_global_color_mapping(unique_project_ids_desvio_hex, style='hex')
                    currency_cols_desvio = ['Valor do Contrato', 'Repasse Recebido', 'Desvio Total (R$)']
                    numeric_cols_desvio = ['Quant.']
                    percentage_cols_desvio = ['% do Projeto', '% Repasse Incorridos', '% Repasse Correlatos']

                    excel_buffer_desvio = create_styled_excel(
                        df_export_desvio,
                        project_id_col=id_column_desvio_hex,
                        color_mapping=hex_color_map_desvio,
                        numeric_cols=numeric_cols_desvio,
                        currency_cols=currency_cols_desvio,
                        percentage_cols=percentage_cols_desvio,
                        filename="desvio_proporcao.xlsx",
                        drop_id_col_on_export=False # Keep ID DO PROJETO for user
                    )
            # If no CLIENTE column or no projects initially flagged, skip all the above processing
            # show_desvio_buttons_and_table remains False

            # 2. Place buttons only if there's deviation content to show/download
            if show_desvio_buttons_and_table:
                st.download_button(
                    label="📥 Download Desvio de Proporção (Excel)",
                    data=excel_buffer_desvio,
                    file_name=f"desvio_proporcao_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                show_table_desvio = st.button("Mostrar Registros com Desvio de Proporção")
                
                # Store display state in session_state to prevent refreshes when other widgets are used
                if show_table_desvio:
                    st.session_state.show_desvio_table = True
                
                # Check if 'show_desvio_table' exists in session state
                if 'show_desvio_table' not in st.session_state:
                    st.session_state.show_desvio_table = False
                    
                # If "Limpar" button is pressed in other sections, we might want to reset this too
                if 'clear_filters' in locals() and clear_filters:
                    st.session_state.show_desvio_table = False

            else:
                 # If no deviations were found after calculation, display message instead of buttons
                 if 'CLIENTE' in projeto_totals.columns: # Check CLIENTE column existed for calc attempt
                      st.info("Não foram encontrados projetos com desvio de proporção significativo.")
                 else:
                      st.warning("Não foi possível processar os desvios: coluna 'CLIENTE' não encontrada nos dados agregados.")
                 st.session_state.show_desvio_table = False # Ensure state is false


            # 3. Display table conditionally
            if st.session_state.show_desvio_table and not df_exibir_desvio.empty: # Check df is not empty
                st.markdown("<h3 style='font-size:140%;'>Planilha de Projetos com Desvio de Proporção</h3>", unsafe_allow_html=True)
                st.info(
                     """
                     *   Destaca projetos com **descompasso financeiro**: a proporção do contrato recebida pela fundação difere significativamente da proporção dos custos (incorridos + correlatos) repassados à Innovatis (considerando uma tolerância).
                     *   **Exemplo:** Se a fundação recebeu 60% do contrato, a Innovatis deveria ter recebido aproximadamente 60% dos custos totais previstos. Um valor muito menor indica um desvio.
                     *   O cálculo é feito **por projeto (em bloco)**, somando todos os recebimentos e custos associados a ele.
                     *   Projetos são agrupados visualmente por cor e **ID DO PROJETO**.
                     """
                 )
                st.info(f"Total de projetos com desvio: {len(df_exibir_desvio)}")

                st.dataframe(
                    df_exibir_desvio.style.format({
                        'Valor do Contrato': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.'),
                        'Repasse Recebido': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.'),
                        'Desvio Total (R$)': lambda x: f"R$ {x:_.2f}".replace('.', ',').replace('_', '.'),
                        '% do Projeto': lambda x: f"{x*100:.1f}%".replace('.', ',') if isinstance(x, (int, float)) and x > 0 else "0,0%",
                        '% Repasse Incorridos': lambda x: f"{x:.1f}%".replace('.', ',') if pd.notna(x) and isinstance(x, (int, float)) and x > 0 else "—",
                        '% Repasse Correlatos': lambda x: f"{x:.1f}%".replace('.', ',') if pd.notna(x) and isinstance(x, (int, float)) and x > 0 else "—",
                        'Quant.': lambda x: f"{int(x)}" if isinstance(x, (int, float)) else "—", # Handle potential non-numeric Quant.
                        'Página': lambda x: str(x) if pd.notna(x) else "—",
                        'Cliente': lambda x: str(x) if pd.notna(x) else "—",
                        'Projeto': lambda x: str(x) if pd.notna(x) else "—",
                        'ID DO PROJETO': lambda x: str(x) if pd.notna(x) else "—"
                    }, na_rep="—").apply(
                        # Pass the final display dataframe and the ID column name
                        lambda s: highlight_projects_detail(s, df_exibir_desvio, 'ID DO PROJETO'),
                        axis=1 # Apply row-wise for background color
                    ),
                    use_container_width=True,
                    hide_index=False  # Exibir explicitamente o índice da linha
                )
                
                # Add a button to close the table
                if st.button("Fechar Tabela", key="btn_fechar_desvio_table"):
                    st.session_state.show_desvio_table = False
                    st.rerun()  # Force rerun to hide the table
            # --- End Desvio de Proporção Section Logic ---

        except Exception as e:
            # Existing error handling logic
            erro_detalhado = f"Erro ao carregar o dashboard de desvios: {str(e)}"
            erro_str = str(e)

            # Verificar se o erro está relacionado a colunas
            if "'CLIENTE'" in erro_str or "not in index" in erro_str or "PAGINA" in erro_str or "'PÁGINA'" in erro_str:
                # Mostrar informações detalhadas sobre as colunas disponíveis
                if 'projeto_totals' in locals():
                    colunas_disponiveis = list(projeto_totals.columns)
                    erro_detalhado += f"\n\nColunas disponíveis em projeto_totals: {colunas_disponiveis}"

                    # Verificar se é problema de nomenclatura (maiúsculas/minúsculas)
                    if "'CLIENTE'" in erro_str:
                        colunas_similares = [col for col in colunas_disponiveis if col.upper() == 'CLIENTE' or 'CLIENT' in col.upper()]
                        if colunas_similares:
                            erro_detalhado += f"\n\nColunas similares a 'CLIENTE' encontradas: {colunas_similares}"

                    # Verificar erros relacionados a PAGINA ou PÁGINA
                    if "'PAGINA'" in erro_str or "'PÁGINA'" in erro_str or "PAGINA" in erro_str:
                        erro_detalhado += "\n\nErro relacionado à coluna PÁGINA. Verifique se a coluna existe e está com o acento correto."

                    # Verificar erros de colunas não encontradas
                    if "not in index" in erro_str:
                        coluna_faltante = erro_str.split("'")[1] if "'" in erro_str else "desconhecida"
                        erro_detalhado += f"\n\nColuna '{coluna_faltante}' não encontrada no DataFrame."

                # Check df_exibir_desvio existence before accessing columns
                if 'df_exibir_desvio' in locals() and isinstance(df_exibir_desvio, pd.DataFrame):
                    erro_detalhado += f"\n\nColunas disponíveis em df_exibir_desvio: {list(df_exibir_desvio.columns)}"

                erro_detalhado += "\n\nVerifique se o nome da coluna está correto ou se os dados foram carregados corretamente."

            st.error(erro_detalhado)

            # Mostrar detalhes do erro para debugging
            if st.checkbox("Mostrar detalhes técnicos para debugging"):
                st.warning("Informações técnicas (para desenvolvedor):")

                # Mostrar informações das colunas
                if 'df_desvio' in locals():
                    st.write("Colunas disponíveis em df_desvio:", list(df_desvio.columns))

                # Mostrar stack trace completo
                import traceback
                st.code(traceback.format_exc())

        # Adicionando espaço após as métricas
        st.markdown("---")

        # Título da Seção de Análise Gráfica
        st.subheader("Análise Gráfica")
        st.markdown("<p style='margin-bottom: 20px;'>Visualizações gráficas dos dados financeiros possivelmente úteis para tomadas de decisão.</p>", unsafe_allow_html=True)

        # Inicializar estado para os filtros dos gráficos se não existir
        if 'graph_filters' not in st.session_state:
            st.session_state.graph_filters = {
                'cliente': {'datas': [], 'tipos': [], 'fundacoes': []},
                'fundacao': {'datas': [], 'tipos': []},
                'tipo': {'datas': [], 'fundacoes': []},
                'custos': {'datas': [], 'fundacoes': [], 'clientes': []}
            }
            
        # Função de callback para atualizar os filtros sem refresh completo
        def update_graph_filter(graph_type, filter_type, value):
            st.session_state.graph_filters[graph_type][filter_type] = value

        # Criando colunas para os gráficos
        row1_col1, row1_col2 = st.columns(2)
        row2_col1, row2_col2 = st.columns(2)

        def gerar_analise_grafico(dados, tipo_grafico):
            """
            Gera análise automática dos dados do gráfico usando IA simples.
            """
            try:
                if tipo_grafico == 'fundacao':
                    if dados.empty:
                        return ""
                    total = dados['SALDO A RECEBER'].sum()
                    maior_fundacao = dados.iloc[0]['FUNDAÇÃO']
                    valor_maior = dados.iloc[0]['SALDO A RECEBER']
                    percentual_maior = (valor_maior / total) * 100 if total > 0 else 0
                    top3_valor = dados.head(3)['SALDO A RECEBER'].sum()
                    top3_percentual = (top3_valor / total) * 100 if total > 0 else 0
                    analise = f"""<div style='background-color: #f0f8ff; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 40px;'><p style='font-size: 14px; margin: 0;'><strong>Análise em tempo real:</strong> A fundação <strong>{maior_fundacao}</strong> representa {percentual_maior:.1f}% do total a receber. As três principais fundações concentram {top3_percentual:.1f}% do valor total.</p></div>"""
                elif tipo_grafico == 'tipo':
                    if dados.empty:
                        return ""
                    total = dados['SALDO A RECEBER'].sum()
                    maior_tipo = dados.iloc[0]['TIPO']
                    valor_maior = dados.iloc[0]['SALDO A RECEBER']
                    percentual_maior = (valor_maior / total) * 100 if total > 0 else 0
                    qtd_tipos = len(dados)
                    diversificacao = "alta" if qtd_tipos >= 4 else "média" if qtd_tipos >= 2 else "baixa"
                    analise = f"""<div style='background-color: #f0f8ff; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 40px;'><p style='font-size: 14px; margin: 0;'><strong>Análise em tempo real:</strong> O serviço <strong>{maior_tipo}</strong> representa {percentual_maior:.1f}% do total a receber. A diversificação de serviços é {diversificacao} com {qtd_tipos} tipos diferentes.</p></div>"""
                elif tipo_grafico == 'cliente':
                    if dados.empty:
                        return ""
                    maior_cliente = dados.iloc[-1]['CLIENTE AGRUPADO']
                    valor_maior = dados.iloc[-1]['SALDO A RECEBER']
                    total = dados['SALDO A RECEBER'].sum()
                    percentual_maior = (valor_maior / total) * 100 if total > 0 else 0
                    tem_outros = 'Outros' in dados['CLIENTE AGRUPADO'].values
                    analise_outros = ""
                    if tem_outros:
                        valor_outros = dados[dados['CLIENTE AGRUPADO'] == 'Outros']['SALDO A RECEBER'].sum()
                        percentual_outros = (valor_outros / total) * 100 if total > 0 else 0
                        analise_outros = f"Clientes menores ('Outros') representam {percentual_outros:.1f}% do faturamento."
                    else:
                        analise_outros = "Não há clientes agrupados na categoria 'Outros'."
                    analise = f"""<div style='background-color: #f0f8ff; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 40px;'><p style='font-size: 14px; margin: 0;'><strong> Análise em tempo real:</strong> O cliente <strong>{maior_cliente}</strong> representa {percentual_maior:.1f}% do valor total a receber. {analise_outros}</p></div>"""
                elif tipo_grafico == 'custos':
                    if not dados or sum(dados) == 0:
                        return ""
                    total = sum(dados)
                    percentual_incorridos = (dados[0] / total) * 100 if total > 0 else 0
                    razao = dados[0] / dados[1] if dados[1] > 0 else float('inf')
                    analise_razao = f"Os custos incorridos são {razao:.1f}x maiores que os correlatos."
                    analise = f"""<div style='background-color: #f0f8ff; padding: 10px; border-radius: 5px; margin-top: 10px; margin-bottom: 40px;'><p style='font-size: 14px; margin: 0;'><strong>Análise em tempo real:</strong> Custos incorridos representam {percentual_incorridos:.1f}% do total. {analise_razao}</p></div>"""
                else:
                    analise = "<div></div>"
                return analise
            except Exception as e:
                return f"<div style='color: #999; font-size: 12px;'>Não foi possível gerar análise: {str(e)}</div>"

        # Gráfico de barras horizontais - Distribuição por Cliente
        with row2_col1:
            st.markdown("<h3 style='font-size: 23px; margin-bottom: 20px;'>Distribuição por Cliente</h3>", unsafe_allow_html=True)
            col_date, col_tipo, col_fundacao = st.columns(3)
            with col_date:
                datas_selecionadas = st.multiselect("Data:", meses_disponiveis_all, default=st.session_state.graph_filters['cliente']['datas'], key="cliente_data_select", on_change=update_graph_filter, args=('cliente', 'datas', st.session_state.get('cliente_data_select', [])))
                st.session_state.graph_filters['cliente']['datas'] = datas_selecionadas
            with col_tipo:
                tipos_disponiveis = sorted(df_desvio['TIPO'].unique())
                tipos_selecionados = st.multiselect("Tipo de Serviço:", tipos_disponiveis, default=st.session_state.graph_filters['cliente']['tipos'], key="cliente_tipo_select", on_change=update_graph_filter, args=('cliente', 'tipos', st.session_state.get('cliente_tipo_select', [])))
                st.session_state.graph_filters['cliente']['tipos'] = tipos_selecionados
            with col_fundacao:
                fundacoes_disponiveis = sorted(df_desvio['FUNDAÇÃO'].unique())
                fundacoes_selecionadas = st.multiselect("Fundação:", fundacoes_disponiveis, default=st.session_state.graph_filters['cliente']['fundacoes'], key="cliente_fundacao_select", on_change=update_graph_filter, args=('cliente', 'fundacoes', st.session_state.get('cliente_fundacao_select', [])))
                st.session_state.graph_filters['cliente']['fundacoes'] = fundacoes_selecionadas
            
            dados_local = df_desvio.copy()
            if st.session_state.graph_filters['cliente']['datas']:
                dados_local = dados_local[dados_local['PREVISÃO DE DATA DE RECEBIMENTO'].isin(st.session_state.graph_filters['cliente']['datas'])]
            if st.session_state.graph_filters['cliente']['tipos']:
                dados_local = dados_local[dados_local['TIPO'].isin(st.session_state.graph_filters['cliente']['tipos'])]
            if st.session_state.graph_filters['cliente']['fundacoes']:
                dados_local = dados_local[dados_local['FUNDAÇÃO'].isin(st.session_state.graph_filters['cliente']['fundacoes'])]
            
            if not dados_local.empty:
                dados_local['SALDO A RECEBER'] = dados_local.groupby('PROJETO_ID_KEY')['SALDO A RECEBER'].transform('first')
                dados_local = dados_local.drop_duplicates(subset=['PROJETO_ID_KEY'])
                dados_local = dados_local[dados_local['SALDO A RECEBER'] > 0]
            
            total_por_cliente = dados_local.groupby('CLIENTE')['SALDO A RECEBER'].sum().reset_index()
            total_por_cliente = total_por_cliente.sort_values(by='SALDO A RECEBER', ascending=False)
            total_por_cliente['CLIENTE AGRUPADO'] = total_por_cliente['CLIENTE']
            total_por_cliente.loc[
                total_por_cliente['SALDO A RECEBER'] / total_por_cliente['SALDO A RECEBER'].sum() < 0.03,
                'CLIENTE AGRUPADO'
            ] = 'Outros'
            
            agrupado = total_por_cliente.groupby('CLIENTE AGRUPADO')['SALDO A RECEBER'].sum().reset_index()
            agrupado = agrupado.sort_values(by='SALDO A RECEBER', ascending=True)
            agrupado['SALDO A RECEBER'] /= 1_000_000
            
            cores = colors_palette[:len(agrupado)]
            fig_bar, ax_bar = plt.subplots(figsize=(3, 2))
            ax_bar.barh(agrupado['CLIENTE AGRUPADO'], agrupado['SALDO A RECEBER'], color=cores)
            ax_bar.set_xlabel('Saldo a Receber (em milhões)', fontsize=5)
            ax_bar.set_ylabel('Cliente', fontsize=5)
            ax_bar.ticklabel_format(style='plain', axis='x', useOffset=False)
            ax_bar.tick_params(axis='x', labelsize=4)
            ax_bar.tick_params(axis='y', labelsize=4)
            for i, v in enumerate(agrupado['SALDO A RECEBER']):
                ax_bar.text(v + (v * 0.01), i, f'R$ {v:,.2f}M'.replace(',', '_').replace('.', ',').replace('_', '.'), va='center', fontsize=4, color='black')
            st.pyplot(fig_bar, use_container_width=False)
            
            # Adicionar análise automática
            analise_cliente = gerar_analise_grafico(agrupado, 'cliente')
            st.markdown(analise_cliente, unsafe_allow_html=True)
            
        # Gráfico de Pizza: Distribuição dos Custos Incorridos e Correlatos
        with row2_col2:
            st.markdown("<h3 style='font-size: 23px; margin-bottom: 20px;'>Distribuição dos Custos</h3>", unsafe_allow_html=True)
            col1, col2, col3 = st.columns(3)
            with col1:
                datas_selecionadas_custos = st.multiselect("Data:", meses_disponiveis_all, default=st.session_state.graph_filters['custos']['datas'], key="custos_data_select", on_change=update_graph_filter, args=('custos', 'datas', st.session_state.get('custos_data_select', [])))
                st.session_state.graph_filters['custos']['datas'] = datas_selecionadas_custos
            with col2:
                fundacoes_disponiveis_custos = sorted(df_desvio['FUNDAÇÃO'].unique())
                fundacoes_selecionadas_custos = st.multiselect("Fundação:", fundacoes_disponiveis_custos, default=st.session_state.graph_filters['custos']['fundacoes'], key="custos_fundacao_select", on_change=update_graph_filter, args=('custos', 'fundacoes', st.session_state.get('custos_fundacao_select', [])))
                st.session_state.graph_filters['custos']['fundacoes'] = fundacoes_selecionadas_custos
            with col3:
                clientes_disponiveis_custos = sorted(df_desvio['CLIENTE'].unique())
                clientes_selecionados_custos = st.multiselect("Cliente:", clientes_disponiveis_custos, default=st.session_state.graph_filters['custos']['clientes'], key="custos_cliente_select", on_change=update_graph_filter, args=('custos', 'clientes', st.session_state.get('custos_cliente_select', [])))
                st.session_state.graph_filters['custos']['clientes'] = clientes_selecionados_custos
                
            dados_local_custos = df_desvio.copy()
            if st.session_state.graph_filters['custos']['datas']:
                dados_local_custos = dados_local_custos[dados_local_custos['PREVISÃO DE DATA DE RECEBIMENTO'].isin(st.session_state.graph_filters['custos']['datas'])]
            if st.session_state.graph_filters['custos']['fundacoes']:
                dados_local_custos = dados_local_custos[dados_local_custos['FUNDAÇÃO'].isin(st.session_state.graph_filters['custos']['fundacoes'])]
            if st.session_state.graph_filters['custos']['clientes']:
                dados_local_custos = dados_local_custos[dados_local_custos['CLIENTE'].isin(st.session_state.graph_filters['custos']['clientes'])]

            total_custos_incurridos, total_custos_correlatos = 0, 0
            if not dados_local_custos.empty:
                unique_project_costs = dados_local_custos.drop_duplicates(subset=['PROJETO_ID_KEY'])
                total_custos_incurridos = unique_project_costs['CUSTOS INCORRIDOS'].sum()
                total_custos_correlatos = unique_project_costs['OUTROS CORRELATOS'].sum()
            custos_labels = ['Custos Incorridos', 'Custos Correlatos']
            custos_values = [total_custos_incurridos, total_custos_correlatos]
            color_map = {
                'Custos Incorridos': '#a1c9f4',
                'Custos Correlatos': '#a1f4c9'
            }
            custos_colors = [color_map[label] for label in custos_labels]
            
            def make_autopct(values):
                def my_autopct(pct):
                    total = sum(values)
                    val = pct * total / 100.0
                    return f'{pct:.1f}%\nR$ {val:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                return my_autopct
            
            fig_pizza, ax_pizza = plt.subplots(figsize=(2, 2.35))
            wedges, texts, autotexts = ax_pizza.pie(
                custos_values,
                labels=custos_labels,
                autopct=make_autopct(custos_values),
                startangle=60,
                colors=custos_colors,
                textprops={'fontsize': 5}
            )
            plt.legend(custos_labels, fontsize=5, loc='center left', bbox_to_anchor=(1, 0.5))
            ax_pizza.axis('equal')
            st.pyplot(fig_pizza, use_container_width=False)
            
            # Adicionar análise automática
            analise_custos = gerar_analise_grafico(custos_values, 'custos')
            st.markdown(analise_custos, unsafe_allow_html=True)
            
        # Gráfico de barras - Distribuição de Valor a Receber por Fundação
        with row1_col1:
            st.markdown("<h3 style='font-size: 23px; margin-bottom: 20px;'>Valor a Receber por Fundação</h3>", unsafe_allow_html=True)
            col_date, col_tipo = st.columns(2)
            with col_date:
                datas_selecionadas_fundacao = st.multiselect("Data:", meses_disponiveis_all, default=st.session_state.graph_filters['fundacao']['datas'], key="fundacao_data_select", on_change=update_graph_filter, args=('fundacao', 'datas', st.session_state.get('fundacao_data_select', [])))
                st.session_state.graph_filters['fundacao']['datas'] = datas_selecionadas_fundacao
            with col_tipo:
                tipos_disponiveis = sorted(df_desvio['TIPO'].unique())
                tipos_selecionados_fund = st.multiselect("Tipo de Serviço:", tipos_disponiveis, default=st.session_state.graph_filters['fundacao']['tipos'], key="fundacao_tipo_select", on_change=update_graph_filter, args=('fundacao', 'tipos', st.session_state.get('fundacao_tipo_select', [])))
                st.session_state.graph_filters['fundacao']['tipos'] = tipos_selecionados_fund
                
            dados_local_fundacao = df_desvio.copy()
            if st.session_state.graph_filters['fundacao']['datas']:
                dados_local_fundacao = dados_local_fundacao[dados_local_fundacao['PREVISÃO DE DATA DE RECEBIMENTO'].isin(st.session_state.graph_filters['fundacao']['datas'])]
            if st.session_state.graph_filters['fundacao']['tipos']:
                dados_local_fundacao = dados_local_fundacao[dados_local_fundacao['TIPO'].isin(st.session_state.graph_filters['fundacao']['tipos'])]
            
            if not dados_local_fundacao.empty:
                dados_local_fundacao['SALDO A RECEBER'] = dados_local_fundacao.groupby('PROJETO_ID_KEY')['SALDO A RECEBER'].transform('first')
                dados_local_fundacao = dados_local_fundacao.drop_duplicates(subset=['PROJETO_ID_KEY'])
                dados_local_fundacao = dados_local_fundacao[dados_local_fundacao['SALDO A RECEBER'] > 0]
            
            total_a_receber_por_fundacao = dados_local_fundacao.groupby('FUNDAÇÃO')['SALDO A RECEBER'].sum().reset_index()
            total_a_receber_por_fundacao['SALDO A RECEBER'] = pd.to_numeric(total_a_receber_por_fundacao['SALDO A RECEBER'], errors='coerce')
            total_a_receber_por_fundacao = total_a_receber_por_fundacao.sort_values(by='SALDO A RECEBER', ascending=False)
            
            fig_bar_fundacao, ax_bar_fundacao = plt.subplots(figsize=(3, 2))
            ax_bar_fundacao.bar(
                total_a_receber_por_fundacao['FUNDAÇÃO'],
                total_a_receber_por_fundacao['SALDO A RECEBER'],
                color=colors_palette[1]
            )
            ax_bar_fundacao.set_ylabel('Valor total a receber', fontsize=5)
            ax_bar_fundacao.set_xlabel('Fundação', fontsize=5)
            for i, v in enumerate(total_a_receber_por_fundacao['SALDO A RECEBER']):
                num_val = float(v)
                ax_bar_fundacao.text(i, num_val + 10000, f'R$ {num_val:,.0f}'.replace(',', '_').replace('.', ',').replace('_', '.'), 
                                    ha='center', va='bottom', fontsize=5)
            plt.ticklabel_format(axis='y', style='plain')
            plt.xticks(rotation=0, ha='center', fontsize=5)
            plt.yticks(fontsize=5)
            plt.tight_layout()
            st.pyplot(fig_bar_fundacao, use_container_width=False)
            
            # Adicionar análise automática
            analise_fundacao = gerar_analise_grafico(total_a_receber_por_fundacao, 'fundacao')
            st.markdown(analise_fundacao, unsafe_allow_html=True)
            
        # Gráfico de barras - Distribuição de Valor a Receber por Tipo de Serviço
        with row1_col2:
            st.markdown("<h3 style='font-size: 23px; margin-bottom: 20px;'>Valor a Receber por Tipo de Serviço</h3>", unsafe_allow_html=True)
            col_date, col_fundacao = st.columns(2)
            with col_date:
                datas_selecionadas_tipo = st.multiselect("Data:", meses_disponiveis_all, default=st.session_state.graph_filters['tipo']['datas'], key="tipo_data_select", on_change=update_graph_filter, args=('tipo', 'datas', st.session_state.get('tipo_data_select', [])))
                st.session_state.graph_filters['tipo']['datas'] = datas_selecionadas_tipo
            with col_fundacao:
                fundacoes_disponiveis_tipo = sorted(df_desvio['FUNDAÇÃO'].unique())
                fundacoes_selecionadas_tipo = st.multiselect("Fundação:", fundacoes_disponiveis_tipo, default=st.session_state.graph_filters['tipo']['fundacoes'], key="tipo_fundacao_select", on_change=update_graph_filter, args=('tipo', 'fundacoes', st.session_state.get('tipo_fundacao_select', [])))
                st.session_state.graph_filters['tipo']['fundacoes'] = fundacoes_selecionadas_tipo
                
            dados_local_tipo = df_desvio.copy()
            if st.session_state.graph_filters['tipo']['datas']:
                dados_local_tipo = dados_local_tipo[dados_local_tipo['PREVISÃO DE DATA DE RECEBIMENTO'].isin(st.session_state.graph_filters['tipo']['datas'])]
            if st.session_state.graph_filters['tipo']['fundacoes']:
                dados_local_tipo = dados_local_tipo[dados_local_tipo['FUNDAÇÃO'].isin(st.session_state.graph_filters['tipo']['fundacoes'])]
            
            if not dados_local_tipo.empty:
                dados_local_tipo['SALDO A RECEBER'] = dados_local_tipo.groupby('PROJETO_ID_KEY')['SALDO A RECEBER'].transform('first')
                dados_local_tipo = dados_local_tipo.drop_duplicates(subset=['PROJETO_ID_KEY'])
                dados_local_tipo = dados_local_tipo[dados_local_tipo['SALDO A RECEBER'] > 0]
            
            total_a_receber_por_tipo = dados_local_tipo.groupby('TIPO')['SALDO A RECEBER'].sum().reset_index()
            total_a_receber_por_tipo['SALDO A RECEBER'] = pd.to_numeric(total_a_receber_por_tipo['SALDO A RECEBER'], errors='coerce')
            total_a_receber_por_tipo = total_a_receber_por_tipo.sort_values(by='SALDO A RECEBER', ascending=False)
            
            fig_bar_tipo, ax_bar_tipo = plt.subplots(figsize=(3, 2))
            ax_bar_tipo.bar(
                total_a_receber_por_tipo['TIPO'],
                total_a_receber_por_tipo['SALDO A RECEBER'],
                color=colors_palette[0]
            )
            ax_bar_tipo.set_ylabel('Valor total a receber', fontsize=5)
            ax_bar_tipo.set_xlabel('Tipo de Serviço', fontsize=5)
            for i, v in enumerate(total_a_receber_por_tipo['SALDO A RECEBER']):
                num_val = float(v)
                ax_bar_tipo.text(i, num_val + 10000, f'R$ {num_val:,.0f}'.replace(',', '_').replace('.', ',').replace('_', '.'), 
                                ha='center', va='bottom', fontsize=5)
            plt.ticklabel_format(axis='y', style='plain')
            plt.xticks(rotation=0, ha='center', fontsize=5)
            plt.yticks(fontsize=5)
            plt.tight_layout()
            st.pyplot(fig_bar_tipo, use_container_width=False)
            
            # Adicionar análise automática
            analise_tipo = gerar_analise_grafico(total_a_receber_por_tipo, 'tipo')
            st.markdown(analise_tipo, unsafe_allow_html=True)
            
        # Rodapé
        st.markdown("---")
        st.markdown("<div style='text-align: center;'>Dashboard Financeiro Versão 1.4 © 2025</div>", unsafe_allow_html=True)

try:
    # Tente executar o dashboard
    pass
except Exception as e:
    # Capture qualquer exceção e exiba para o usuário
    st.error(f"Ocorreu um erro inesperado na execução do script: {str(e)}")
